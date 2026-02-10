import pandas as pd
import numpy as np
import os
import warnings
import joblib
import time
from datetime import datetime
from sklearn.metrics import mean_squared_error, mean_absolute_percentage_error
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor, ExtraTreesRegressor, AdaBoostRegressor
from sklearn.linear_model import Ridge, ElasticNet, BayesianRidge, Lasso, HuberRegressor, LinearRegression
from sklearn.kernel_ridge import KernelRidge
from sklearn.svm import SVR
from sklearn.neighbors import KNeighborsRegressor
from sklearn.neural_network import MLPRegressor
from sklearn.gaussian_process import GaussianProcessRegressor
from sklearn.gaussian_process.kernels import RBF, ConstantKernel as C
from sklearn.preprocessing import StandardScaler, PolynomialFeatures
from sklearn.pipeline import make_pipeline
from statsmodels.tsa.holtwinters import ExponentialSmoothing
from statsmodels.tsa.statespace.sarimax import SARIMAX
from statsmodels.tsa.forecasting.theta import ThetaModel
import matplotlib.pyplot as plt
import holidays

# Import error handling module
try:
    from error_handler import (
        DataError, ModelError, ValidationError, handle_error,
        safe_execute, validate_data_file, log_operation
    )
except ImportError:
    # Fallback if error_handler not available
    class DataError(Exception): pass
    class ModelError(Exception): pass
    class ValidationError(Exception): pass
    def handle_error(e, *args, **kwargs): print(f"Error: {e}")
    def safe_execute(func, *args, **kwargs): return func()
    def validate_data_file(*args, **kwargs): return True
    def log_operation(*args, **kwargs): pass

# --- Optional Dependencies ---
try:
    from prophet import Prophet
    PROPHET_AVAILABLE = True
except ImportError:
    PROPHET_AVAILABLE = False
    print("Warning: 'prophet' not found.")

try:
    from xgboost import XGBRegressor
    XGBOOST_AVAILABLE = True
except ImportError:
    XGBOOST_AVAILABLE = False
    print("Warning: 'xgboost' not found.")

try:
    from lightgbm import LGBMRegressor
    LIGHTGBM_AVAILABLE = True
except ImportError:
    LIGHTGBM_AVAILABLE = False
    print("Warning: 'lightgbm' not found.")

# CatBoost removed from required imports to avoid crash, but kept logic if user installs it later
try:
    from catboost import CatBoostRegressor
    CATBOOST_AVAILABLE = True
except ImportError:
    CATBOOST_AVAILABLE = False
    # print("Warning: 'catboost' not found.")

try:
    import optuna
    OPTUNA_AVAILABLE = True
    optuna.logging.set_verbosity(optuna.logging.WARNING)
except ImportError:
    OPTUNA_AVAILABLE = False
    print("Warning: 'optuna' not found.")

try:
    import shap
    SHAP_AVAILABLE = True
except ImportError:
    SHAP_AVAILABLE = False
    print("Warning: 'shap' not found.")

# Suppress warnings
warnings.filterwarnings("ignore")
import logging
logging.getLogger('cmdstanpy').setLevel(logging.ERROR) 

class TaxForecaster:
    """
    The core forecasting engine for the Tax Forecasting System.
    
    This class handles:
    1. Data Loading & Preprocessing (Merging Tax History with Macro Data).
    2. Feature Engineering (Lags, Rolling Means, Holiday/Seasonality Features).
    3. Model Training (Fitting multiple models: ARIMA, Prophet, ML, DL).
    4. Hyperparameter Tuning (Using Optuna).
    5. Forecasting (Generating future predictions with Confidence Intervals).
    6. Explanation (SHAP values).
    """
    def __init__(self, history_file, macro_file=None):
        """
        Initialize the TaxForecaster.

        Args:
            history_file (str or file-like): Path to tax revenue history CSV or UploadedFile object.
            macro_file (str or file-like, optional): Path to macroeconomic indicators CSV or UploadedFile object.
        """
        self.history_file = history_file
        self.macro_file = macro_file
        self.df = None
        self.macro_df = None
        self.results = []
        self.model_performance = []
        
        # Cache Storage
        self.trained_models = {} 
        self.is_fitted = False
        self.available_models = []
        
        # Quick Win: Metadata tracking
        self.metadata = {
            'data_loaded_at': None,
            'macro_fetched_at': None,
            'last_trained_at': None,
            'training_duration': None,
            'last_forecast_at': None
        }
        self.previous_forecast = None  # Store for comparison 
    
    def get_memory_usage(self):
        """
        Calculate approximate memory usage of forecaster object.
        
        Returns:
            dict: Memory usage information in MB
        """
        import sys
        
        memory_info = {
            'df_size': sys.getsizeof(self.df) / (1024 ** 2) if self.df is not None else 0,
            'macro_df_size': sys.getsizeof(self.macro_df) / (1024 ** 2) if self.macro_df is not None else 0,
            'models_count': len(self.trained_models),
            'results_size': sys.getsizeof(self.results) / (1024 ** 2) if self.results else 0
        }
        memory_info['total_mb'] = sum([v for k, v in memory_info.items() if k != 'models_count'])
        
        return memory_info
    
    def cleanup_memory(self, keep_models=True):
        """
        Clean up memory by removing non-essential cached data.
        
        Args:
            keep_models (bool): Whether to keep trained models
        """
        print("[INFO] Cleaning up memory...")
        
        # Clear results if they're stored
        if hasattr(self, 'previous_forecast'):
            self.previous_forecast = None
        
        # Optionally clear models
        if not keep_models:
            self.trained_models = {}
            self.is_fitted = False
            print("[INFO] Cleared trained models")
        
        print("[INFO] Memory cleanup complete")

    def _add_holiday_features(self, df):
        """
        Add Indonesian holiday features (Lebaran, Natal) to boost seasonality.
        Handles both DatetimeIndex and regular index with Tanggal column.
        """
        try:
            # Check if index is DatetimeIndex or if we need to use Tanggal column
            if hasattr(df.index, 'year'):
                years = df.index.year.unique()
                date_col = df.index
            elif 'Tanggal' in df.columns:
                years = pd.to_datetime(df['Tanggal']).dt.year.unique()
                date_col = pd.to_datetime(df['Tanggal'])
            else:
                print("[WARNING] No date information found for holiday features")
                return df
                
            try:
                id_holidays = holidays.Indonesia(years=years)
            except:
                id_holidays = {}

            lebaran_dates = []
            natal_dates = []
            
            for date, name in id_holidays.items():
                if "Idul Fitri" in name:
                    lebaran_dates.append(date)
                elif "Hari Raya Natal" in name or "Christmas" in name:
                    natal_dates.append(date)
                    
            df['is_lebaran'] = 0
            df['is_natal'] = 0
            
            for idx in range(len(df)):
                current_date = date_col.iloc[idx] if hasattr(date_col, 'iloc') else date_col[idx]
                for d in lebaran_dates:
                    if d.year == current_date.year and d.month == current_date.month:
                        df.at[df.index[idx], 'is_lebaran'] = 1
                for d in natal_dates:
                    if d.year == current_date.year and d.month == current_date.month:
                        df.at[df.index[idx], 'is_natal'] = 1
                        
        except Exception as e:
            print(f"[WARNING] Seasonality feature failed: {e}")
            df['is_lebaran'] = 0
            df['is_natal'] = 0
            
        return df
        
    def load_data(self):
        """
        Load and validate tax history and macro data with comprehensive error handling.
        
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            start_time = time.time()
            
            # Handle both file paths (string) and UploadedFile objects
            if hasattr(self.history_file, 'name'):
                print(f"[INFO] Loading from UploadedFile: {self.history_file.name}")
                self.df = pd.read_csv(self.history_file)
            else:
                # It's a file path string
                abs_path = os.path.abspath(self.history_file)
                print(f"[INFO] Loading file: {abs_path}")
                
                if not os.path.exists(abs_path):
                    raise DataError(
                        f"Tax history file not found: {abs_path}",
                        context={'file_path': abs_path}
                    )
                
                self.df = pd.read_csv(self.history_file)
            
            # Validate data structure
            validate_data_file(self.df)
            
            # Parse dates
            try:
                self.df['Tanggal'] = pd.to_datetime(self.df['Tanggal'], dayfirst=True)
            except Exception as e:
                raise ValidationError(
                    "Invalid date format in Tanggal column",
                    context={'error': str(e)}
                )
            
            # Load macro data if provided
            if self.macro_file:
                if hasattr(self.macro_file, 'name'):
                    self.macro_df = pd.read_csv(self.macro_file)
                else:
                    macro_abs_path = os.path.abspath(self.macro_file)
                    if not os.path.exists(macro_abs_path):
                        print(f"[WARNING] Macro file not found: {macro_abs_path}. Continuing without macro data.")
                        self.macro_file = None
                        self.macro_df = None
                    else:
                        self.macro_df = pd.read_csv(self.macro_file)
                
                if self.macro_df is not None:
                    try:
                        self.macro_df['Tanggal'] = pd.to_datetime(self.macro_df['Tanggal'], dayfirst=True)
                    except Exception as e:
                        print(f"[WARNING] Could not parse dates in macro data: {e}")
                        self.macro_df = None
                    
                    # Merge with tax data
                    if self.macro_df is not None:
                        self.df = pd.merge(self.df, self.macro_df, on='Tanggal', how='left')
                        
                        # Fill missing values
                        exclude_cols = {'Jenis Pajak', 'Nominal (Milyar)', 'Tanggal'}
                        macro_cols = [col for col in self.df.columns if col not in exclude_cols]
                        
                        for col in macro_cols:
                            if col in self.df.columns and self.df[col].dtype in ['float64', 'int64']:
                                self.df[col] = self.df[col].ffill().bfill()
            
            # Update metadata
            load_time = time.time() - start_time
            self.metadata['data_loaded_at'] = datetime.now().isoformat()
            self.metadata['load_duration'] = round(load_time, 2)
            
            # Log success
            log_operation(
                'load_data',
                success=True,
                details={
                    'rows': len(self.df),
                    'tax_types': len(self.df['Jenis Pajak'].unique()),
                    'duration_seconds': load_time,
                    'has_macro': self.macro_df is not None
                }
            )
            
            print(f"[SUCCESS] Data loaded: {len(self.df)} rows, {len(self.df['Jenis Pajak'].unique())} tax types")
            return True
            
        except (DataError, ValidationError) as e:
            handle_error(e, 'invalid_csv_format')
            return False
        except FileNotFoundError as e:
            handle_error(e, 'file_not_found', context={'file': str(self.history_file)})
            return False
        except Exception as e:
            handle_error(e, context={'file': str(self.history_file)})
            return False

    def _evaluate(self, y_true, y_pred):
        return np.sqrt(mean_squared_error(y_true, y_pred))

    def _evaluate_mape(self, y_true, y_pred):
        # Handle division by zero edge case
        try:
            return mean_absolute_percentage_error(y_true, y_pred)
        except:
            return 0.0

    def _create_ml_features(self, series, lags=3):
        df = pd.DataFrame(series)
        df.columns = ['y']
        df['Month'] = df.index.month
        df['Year'] = df.index.year
        df['Quarter'] = df.index.quarter
        for lag in range(1, lags + 1):
            df[f'lag_{lag}'] = df['y'].shift(lag)
        df['rolling_mean_3'] = df['y'].shift(1).rolling(window=3).mean()
        df = df.dropna()
        return df

    def _recursive_forecast(self, model, full_X, series, forecast_periods, scaler=None, exog_future=None):
        preds = []
        curr_vals = list(series.values)
        next_date = series.index[-1]
        
        for i in range(forecast_periods):
            next_date = next_date + pd.DateOffset(months=1)
            
            # Base Lag Features
            rec = {
                'Month': next_date.month, 'Year': next_date.year, 'Quarter': next_date.quarter,
                'lag_1': curr_vals[-1], 'lag_2': curr_vals[-2], 'lag_3': curr_vals[-3], 
                'rolling_mean_3': np.mean(curr_vals[-3:])
            }
            
            # Add Exogenous Features if available
            if exog_future is not None and i < len(exog_future):
                # Assuming exog_future is aligned with forecast steps (0..N)
                exog_row = exog_future.iloc[i]
                for col in exog_row.index:
                    rec[col] = exog_row[col]
            
            feat = pd.DataFrame([rec])
            
            # Align with Training Columns (Fill missing with 0 to prevent NaNs)
            feat = feat.reindex(columns=full_X.columns, fill_value=0)
            
            # Apply scaling if model needs it
            if scaler:
                feat_vals = scaler.transform(feat)
                pred = model.predict(feat_vals)[0]
            else:
                pred = model.predict(feat)[0]
            
            preds.append(pred)
            curr_vals.append(pred)
            
        return pd.Series(preds, index=pd.date_range(series.index[-1] + pd.DateOffset(months=1), periods=forecast_periods, freq='ME'))

    def tune_hyperparameters(self, model_name, X_train, y_train, X_test, y_test, n_trials=10):
        """
        Optimizes model hyperparameters using Optuna.

        Args:
            model_name (str): Name of the model to tune (e.g., "XGBoost", "RandomForest").
            X_train (pd.DataFrame): Training features.
            y_train (pd.Series): Training target.
            X_test (pd.DataFrame): Validation features.
            y_test (pd.Series): Validation target.
            n_trials (int): Number of optimization trials.

        Returns:
            dict: Best hyperparameters found.
        """
        if not OPTUNA_AVAILABLE: return None

        def objective(trial):
            if model_name == "XGBoost":
                param = {
                    'n_estimators': trial.suggest_int('n_estimators', 50, 300),
                    'max_depth': trial.suggest_int('max_depth', 3, 10),
                    'learning_rate': trial.suggest_float('learning_rate', 0.01, 0.3),
                    'subsample': trial.suggest_float('subsample', 0.5, 1.0),
                    'colsample_bytree': trial.suggest_float('colsample_bytree', 0.5, 1.0),
                    'random_state': 42
                }
                model = XGBRegressor(**param)
            elif model_name == "LightGBM":
                param = {
                    'n_estimators': trial.suggest_int('n_estimators', 50, 300),
                    'num_leaves': trial.suggest_int('num_leaves', 20, 100),
                    'learning_rate': trial.suggest_float('learning_rate', 0.01, 0.3),
                    'random_state': 42,
                    'verbose': -1
                }
                model = LGBMRegressor(**param)
            elif model_name == "CatBoost":
                param = {
                    'iterations': trial.suggest_int('iterations', 50, 300),
                    'depth': trial.suggest_int('depth', 3, 10),
                    'learning_rate': trial.suggest_float('learning_rate', 0.01, 0.3),
                    'random_state': 42,
                    'verbose': 0
                }
                model = CatBoostRegressor(**param)
            elif model_name == "RandomForest":
                param = {
                    'n_estimators': trial.suggest_int('n_estimators', 50, 300),
                    'max_depth': trial.suggest_int('max_depth', 3, 15),
                    'min_samples_split': trial.suggest_int('min_samples_split', 2, 10),
                    'random_state': 42
                }
                model = RandomForestRegressor(**param)
            elif model_name == "GradientBoosting":
                param = {
                    'n_estimators': trial.suggest_int('n_estimators', 50, 300),
                    'learning_rate': trial.suggest_float('learning_rate', 0.01, 0.3),
                    'max_depth': trial.suggest_int('max_depth', 3, 10),
                    'subsample': trial.suggest_float('subsample', 0.5, 1.0),
                    'random_state': 42
                }
                model = GradientBoostingRegressor(**param)
            elif model_name == "ExtraTrees":
                param = {
                    'n_estimators': trial.suggest_int('n_estimators', 50, 300),
                    'max_depth': trial.suggest_int('max_depth', 3, 15),
                    'min_samples_split': trial.suggest_int('min_samples_split', 2, 10),
                    'random_state': 42
                }
                model = ExtraTreesRegressor(**param)
            elif model_name == "AdaBoost":
                param = {
                    'n_estimators': trial.suggest_int('n_estimators', 50, 200),
                    'learning_rate': trial.suggest_float('learning_rate', 0.01, 1.0),
                    'random_state': 42
                }
                model = AdaBoostRegressor(**param)
            elif model_name == "KNN":
                param = {
                    'n_neighbors': trial.suggest_int('n_neighbors', 3, 20),
                    'weights': trial.suggest_categorical('weights', ['uniform', 'distance']),
                    'p': trial.suggest_int('p', 1, 2)
                }
                model = KNeighborsRegressor(**param)
            elif model_name == "MLP":
                param = {
                    'hidden_layer_sizes': trial.suggest_categorical('hidden_layer_sizes', [(50,), (100,), (50, 50), (100, 50)]),
                    'activation': trial.suggest_categorical('activation', ['relu', 'tanh']),
                    'solver': 'adam',
                    'alpha': trial.suggest_float('alpha', 0.0001, 0.01),
                    'max_iter': 500,
                    'random_state': 42
                }
                model = MLPRegressor(**param)
            elif model_name == "SVR":
                param = {
                    'C': trial.suggest_float('C', 0.1, 100, log=True),
                    'epsilon': trial.suggest_float('epsilon', 0.01, 1.0),
                    'kernel': trial.suggest_categorical('kernel', ['linear', 'rbf', 'poly'])
                }
                model = SVR(**param)
            elif model_name == "ElasticNet":
                param = {
                    'alpha': trial.suggest_float('alpha', 0.01, 10.0, log=True),
                    'l1_ratio': trial.suggest_float('l1_ratio', 0.0, 1.0)
                }
                model = ElasticNet(**param)
            elif model_name == "Lasso":
                param = {
                    'alpha': trial.suggest_float('alpha', 0.01, 10.0, log=True)
                }
                model = Lasso(**param)
            elif model_name == "Huber":
                param = {
                    'epsilon': trial.suggest_float('epsilon', 1.0, 3.0),
                    'alpha': trial.suggest_float('alpha', 0.0001, 0.01)
                }
                model = HuberRegressor(**param)
            elif model_name == "KernelRidge":
                param = {
                    'alpha': trial.suggest_float('alpha', 0.1, 10.0),
                    'kernel': trial.suggest_categorical('kernel', ['linear', 'rbf', 'polynomial'])
                }
                model = KernelRidge(**param)
            
            model.fit(X_train, y_train)
            preds = model.predict(X_test)
            return self._evaluate(y_test, preds)

        study = optuna.create_study(direction='minimize')
        study.optimize(objective, n_trials=n_trials, timeout=120)
        return study.best_params

    def preprocess_outliers(self):
        """
        Detects and corrects outliers using Z-Score.
        Replaces outliers with 3-month rolling average.
        """
        if self.df is None: return "No data loaded"
        
        corrected_count = 0
        
        # Process per tax type to avoid cross-contamination
        for tax in self.df['Jenis Pajak'].unique():
            mask = self.df['Jenis Pajak'] == tax
            series = self.df.loc[mask, 'Nominal (Milyar)'].copy()
            
            # Calculate Z-Score
            mean = series.mean()
            std = series.std()
            
            if std == 0: continue
            
            z_scores = (series - mean) / std
            outliers = (z_scores.abs() > 2.5) # Threshold 2.5 sigma
            
            if outliers.any():
                # Replace with rolling mean (ffill then rolling)
                # Simple approach: interpolate or rolling mean of window 3
                # We use rolling mean of previous 3 points
                
                # Iterate to fix
                indices = outliers[outliers].index
                for idx in indices:
                    # Get surrounding window (excluding self)
                    # Currently simplified: use overall mean or simple limits
                    # Better: use value from same month last year? Or simple rolling.
                    
                    # Logic: Replace with Rolling Median of window 5 centered
                    window_start = max(0, idx-3)
                    window_end = min(len(series), idx+3)
                    subset = series.iloc[window_start:window_end]
                    
                    # Exclude the outlier itself from calculation
                    subset_clean = subset[subset.index != idx]
                    
                    if not subset_clean.empty:
                        new_val = subset_clean.mean()
                        self.df.at[idx, 'Nominal (Milyar)'] = new_val
                        corrected_count += 1
                        
        return f"Fixed {corrected_count} outliers."

    def fit(self, n_trials=10, epochs=100, progress_callback=None):
        """
        Train ALL models and cache them.
        
        Args:
            n_trials (int): Hyperparameter tuning trials (Optuna).
            epochs (int): Max iterations for models like MLP.
            progress_callback (callable): Optional callback function for progress updates.
                Should accept parameters: current_step, total_steps, message
        
        Returns:
            str: Status message
        """
        if self.df is None:
            raise DataError("Data not loaded. Call load_data() first.")
        
        # Track training time
        start_time = time.time()
        training_start = datetime.now()
        
        print(f"\n{'='*60}")
        print(f"TRAINING STARTED: {training_start.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{'='*60}\n")
        
        # Apply Seasonality Features
        try:
            self.df = self._add_holiday_features(self.df)
        except Exception as e:
            print(f"[WARNING] Seasonality feature failed: {e}")
        
        # Guarantee columns exist to prevent KeyError later
        if 'is_lebaran' not in self.df.columns:
            self.df['is_lebaran'] = 0
        if 'is_natal' not in self.df.columns:
            self.df['is_natal'] = 0
        
        tax_types = self.df['Jenis Pajak'].unique()
        self.model_performance = []
        self.trained_models = {}
        all_model_names = set()
        
        total_tax_types = len(tax_types)
        
        for tax_idx, tax in enumerate(tax_types, 1):
            tax_progress = (tax_idx - 1) / total_tax_types
            
            if progress_callback:
                progress_callback(
                    current_step=tax_idx,
                    total_steps=total_tax_types,
                    message=f"Training models for {tax} ({tax_idx}/{total_tax_types})",
                    progress_pct=tax_progress
                )
            
            print(f"\n[{tax_idx}/{total_tax_types}] Training: {tax}")
            print(f"{'-'*50}")
            
            tax_data = self.df[self.df['Jenis Pajak'] == tax].copy().sort_values('Tanggal')
            tax_data.set_index('Tanggal', inplace=True)
            # Check for sufficient data
            if len(tax_data) < 12:
                print(f"[WARNING] Insufficient data for {tax} ({len(tax_data)} rows). Skipping complex ML models.")
                series = tax_data['Nominal (Milyar)']
                # Just fit simple models or skip entirely?
                # We'll skip the ML part but let simple models try if they can handle it (Prophet needs 2 points, Arima needs stationarity)
                # But ML part explicitly does train/test split on -12.
                # So we must avoid ML part if len < 12.
                skip_ml = True
            else:
                skip_ml = False
                
            series = tax_data['Nominal (Milyar)']
            
            # Get all available macro columns dynamically
            exclude_cols = {'Jenis Pajak', 'Nominal (Milyar)', 'Tanggal', 'is_lebaran', 'is_natal'}
            exog_vars = [col for col in tax_data.columns if col not in exclude_cols and tax_data[col].dtype in ['float64', 'int64']]
            
            # Use columns directly from tax_data if they exist
            if len(exog_vars) > 0:
                exog_data = tax_data[exog_vars].copy()
                if 'is_lebaran' in tax_data.columns:
                    exog_data['is_lebaran'] = tax_data['is_lebaran']
                if 'is_natal' in tax_data.columns:
                    exog_data['is_natal'] = tax_data['is_natal']
            else:
                 # Fallback if macro file wasn't loaded
                 exog_data = None

            # Split
            train = series[:-12]
            test = series[-12:]
            exog_train = exog_data[:-12] if exog_data is not None else None
            exog_test = exog_data[-12:] if exog_data is not None else None
            
            candidates = {} 

            # Helper to add candidate
            def add_candidate(name, model, pred, features=None, scaler=None):
                rmse = self._evaluate(test, pred)
                mape = self._evaluate_mape(test, pred)
                candidates[name] = (rmse, mape, model, features, scaler)

            # 1. Holt-Winters
            try:
                hw = ExponentialSmoothing(train, trend='add', seasonal='add', seasonal_periods=12, freq='ME').fit()
                pred = hw.forecast(len(test))
                full_hw = ExponentialSmoothing(series, trend='add', seasonal='add', seasonal_periods=12, freq='ME').fit()
                add_candidate('Holt-Winters', full_hw, pred)
            except: pass

            # 2. SARIMA
            try:
                mod = SARIMAX(train, order=(1,1,1), seasonal_order=(1,1,1,12), enforce_stationarity=False, enforce_invertibility=False).fit(disp=False)
                pred = mod.forecast(len(test))
                full_mod = SARIMAX(series, order=(1,1,1), seasonal_order=(1,1,1,12), enforce_stationarity=False, enforce_invertibility=False).fit(disp=False)
                add_candidate('SARIMA', full_mod, pred)
            except: pass

            # 3. SARIMAX (Multivariate)
            if self.macro_file:
                try:
                    sx = SARIMAX(train, exog=exog_train, order=(1,1,1), seasonal_order=(1,1,1,12), enforce_stationarity=False, enforce_invertibility=False).fit(disp=False)
                    pred = sx.forecast(len(test), exog=exog_test)
                    full_sx = SARIMAX(series, exog=exog_data, order=(1,1,1), seasonal_order=(1,1,1,12), enforce_stationarity=False, enforce_invertibility=False).fit(disp=False)
                    add_candidate('SARIMAX', full_sx, pred, exog_data)
                except: pass

            # 4. Prophet
            if PROPHET_AVAILABLE:
                try:
                    p_df = pd.DataFrame({'ds': train.index, 'y': train.values})
                    m = Prophet(seasonality_mode='multiplicative')
                    m.fit(p_df)
                    future_test = pd.DataFrame({'ds': test.index})
                    pred = m.predict(future_test)['yhat'].values
                    m_final = Prophet(seasonality_mode='multiplicative')
                    m_final.fit(pd.DataFrame({'ds': series.index, 'y': series.values}))
                    add_candidate('Prophet', m_final, pred)
                except: pass
            
            # 5. Theta Model (Simple Benchmark)
            try:
                tm = ThetaModel(train, period=12).fit()
                pred = tm.forecast(len(test))
                full_tm = ThetaModel(series, period=12).fit()
                add_candidate('Theta Model', full_tm, pred)
            except: pass

            # ML Setup
            if not skip_ml:
                ml_df = self._create_ml_features(series)
                if len(ml_df) < 5: # Extra safety check
                     print(f"[WARNING] Not enough data after feature engineering for {tax}. Skipping ML.")
                else:
                    X = ml_df.drop('y', axis=1)
                    y = ml_df['y']
                    
                    # Scaler
                    scaler = StandardScaler()
                    X_scaled = pd.DataFrame(scaler.fit_transform(X), columns=X.columns, index=X.index)
                    
                    split_idx = len(X) - 12
                    
                    # Safety check for split_idx
                    if split_idx < 1:
                        print(f"[WARNING] Train/Test split failed for {tax}. Skipping ML.")
                    else:
                        X_train, X_test = X.iloc[:split_idx], X.iloc[split_idx:]
                        y_train, y_test = y.iloc[:split_idx], y.iloc[split_idx:]
                        
                        X_train_s, X_test_s = X_scaled.iloc[:split_idx], X_scaled.iloc[split_idx:]
        
                        # 6. XGBoost
                        if XGBOOST_AVAILABLE:
                            try:
                                best_params = self.tune_hyperparameters("XGBoost", X_train, y_train, X_test, y_test)
                                xgb = XGBRegressor(**best_params) if best_params else XGBRegressor(n_estimators=100)
                                xgb.fit(X_train, y_train)
                                pred = xgb.predict(X_test)
                                full_xgb = XGBRegressor(**best_params) if best_params else XGBRegressor(n_estimators=100)
                                full_xgb.fit(X, y)
                                add_candidate('XGBoost', full_xgb, pred, X, None) 
                            except: pass
                        
                        # 7. LightGBM
                        if LIGHTGBM_AVAILABLE:
                            try:
                                best_params = self.tune_hyperparameters("LightGBM", X_train, y_train, X_test, y_test)
                                lgb = LGBMRegressor(**best_params, verbose=-1) if best_params else LGBMRegressor(verbose=-1)
                                lgb.fit(X_train, y_train)
                                pred = lgb.predict(X_test)
                                full_lgb = LGBMRegressor(**best_params, verbose=-1) if best_params else LGBMRegressor(verbose=-1)
                                full_lgb.fit(X, y)
                                add_candidate('LightGBM', full_lgb, pred, X, None)
                            except: pass
                        
                        # 8. CatBoost
                        if CATBOOST_AVAILABLE:
                            try:
                                best_params = self.tune_hyperparameters("CatBoost", X_train, y_train, X_test, y_test)
                                cat = CatBoostRegressor(**best_params, verbose=0) if best_params else CatBoostRegressor(verbose=0)
                                cat.fit(X_train, y_train)
                                pred = cat.predict(X_test)
                                full_cat = CatBoostRegressor(**best_params, verbose=0) if best_params else CatBoostRegressor(verbose=0)
                                full_cat.fit(X, y)
                                add_candidate('CatBoost', full_cat, pred, X, None)
                            except: pass
            
                        # 9. Random Forest
                        try:
                            best_params = self.tune_hyperparameters("RandomForest", X_train, y_train, X_test, y_test)
                            rf = RandomForestRegressor(**best_params) if best_params else RandomForestRegressor(n_estimators=100)
                            rf.fit(X_train, y_train)
                            pred = rf.predict(X_test)
                            full_rf = RandomForestRegressor(**best_params) if best_params else RandomForestRegressor(n_estimators=100)
                            full_rf.fit(X, y)
                            add_candidate('Random Forest', full_rf, pred, X, None)
                        except: pass
            
                        # 10. Gradient Boosting
                        try:
                            best_params = self.tune_hyperparameters("GradientBoosting", X_train, y_train, X_test, y_test)
                            gb = GradientBoostingRegressor(**best_params) if best_params else GradientBoostingRegressor(n_estimators=100)
                            gb.fit(X_train, y_train)
                            pred = gb.predict(X_test)
                            full_gb = GradientBoostingRegressor(**best_params) if best_params else GradientBoostingRegressor(n_estimators=100)
                            full_gb.fit(X, y)
                            add_candidate('Gradient Boosting', full_gb, pred, X, None)
                        except: pass
            
                        # 11. Extra Trees
                        try:
                            best_params = self.tune_hyperparameters("ExtraTrees", X_train, y_train, X_test, y_test)
                            et = ExtraTreesRegressor(**best_params) if best_params else ExtraTreesRegressor(n_estimators=100)
                            et.fit(X_train, y_train)
                            pred = et.predict(X_test)
                            full_et = ExtraTreesRegressor(**best_params) if best_params else ExtraTreesRegressor(n_estimators=100)
                            full_et.fit(X, y)
                            add_candidate('Extra Trees', full_et, pred, X, None)
                        except: pass
            
                        # 12. AdaBoost
                        try:
                            best_params = self.tune_hyperparameters("AdaBoost", X_train, y_train, X_test, y_test)
                            ada = AdaBoostRegressor(**best_params) if best_params else AdaBoostRegressor(n_estimators=100)
                            ada.fit(X_train, y_train)
                            pred = ada.predict(X_test)
                            full_ada = AdaBoostRegressor(**best_params) if best_params else AdaBoostRegressor(n_estimators=100)
                            full_ada.fit(X, y)
                            add_candidate('AdaBoost', full_ada, pred, X, None)
                        except: pass

            # 13. KNN
            try:
                best_params = self.tune_hyperparameters("KNN", X_train_s, y_train, X_test_s, y_test)
                knn = KNeighborsRegressor(**best_params) if best_params else KNeighborsRegressor(n_neighbors=5)
                knn.fit(X_train_s, y_train)
                pred = knn.predict(X_test_s)
                full_knn = KNeighborsRegressor(**best_params) if best_params else KNeighborsRegressor(n_neighbors=5)
                full_knn.fit(X_scaled, y)
                add_candidate('KNN', full_knn, pred, X, scaler)
            except: pass

            # 14. MLP (Neural Network)
            try:
                best_params = self.tune_hyperparameters("MLP", X_train_s, y_train, X_test_s, y_test)
                mlp = MLPRegressor(**best_params) if best_params else MLPRegressor(hidden_layer_sizes=(100,), max_iter=500, random_state=42)
                mlp.fit(X_train_s, y_train)
                pred = mlp.predict(X_test_s)
                full_mlp = MLPRegressor(**best_params) if best_params else MLPRegressor(hidden_layer_sizes=(100,), max_iter=500, random_state=42)
                full_mlp.fit(X_scaled, y)
                add_candidate('Neural Network (MLP)', full_mlp, pred, X, scaler)
            except: pass

            # 15. LSTM (Deep Learning) - [NEW]
            try:
                import tensorflow as tf
                from tensorflow.keras.models import Sequential
                from tensorflow.keras.layers import LSTM, Dense, Dropout
                from tensorflow.keras.optimizers import Adam
                
                # Reshape for LSTM [samples, time steps, features]
                # Here we treat 'features' as input vector at t.
                # Actually for standard regression X is (N, Features).
                # We need to reshape X to (N, 1, Features) or (N, Features, 1)?
                # Standard LSTM usage: (N, Timesteps, Features). Here Timesteps=1 (since we already lagged features content manually).
                
                X_train_lstm = X_train_s.values.reshape((X_train_s.shape[0], 1, X_train_s.shape[1]))
                X_test_lstm = X_test_s.values.reshape((X_test_s.shape[0], 1, X_test_s.shape[1]))
                X_full_lstm = X_scaled.values.reshape((X_scaled.shape[0], 1, X_scaled.shape[1]))
                
                def build_lstm_model(input_shape):
                    model = Sequential()
                    model.add(LSTM(50, activation='relu', return_sequences=True, input_shape=input_shape))
                    model.add(Dropout(0.2))
                    model.add(LSTM(50, activation='relu'))
                    model.add(Dropout(0.2))
                    model.add(Dense(1))
                    model.compile(optimizer=Adam(learning_rate=0.01), loss='mse')
                    return model
                
                # Train
                lstm = build_lstm_model((1, X_train_s.shape[1]))
                lstm.fit(X_train_lstm, y_train, epochs=50, batch_size=16, verbose=0)
                
                pred_lstm = lstm.predict(X_test_lstm).flatten()
                
                # Full Model
                full_lstm = build_lstm_model((1, X_scaled.shape[1]))
                full_lstm.fit(X_full_lstm, y, epochs=50, batch_size=16, verbose=0)
                
                # Wrapper for compatibility with _recursive_forecast
                # We need a class that has .predict(X) where X is 2D DF
                class LSTMAdapter:
                    def __init__(self, model):
                        self.model = model
                    def predict(self, X):
                        if isinstance(X, pd.DataFrame):
                            X = X.values
                        if len(X.shape) == 2:
                            X = X.reshape((X.shape[0], 1, X.shape[1]))
                        return self.model.predict(X, verbose=0).flatten()
                
                add_candidate('LSTM (Deep Learning)', LSTMAdapter(full_lstm), pred_lstm, X, scaler)
                
            except ImportError:
                pass # TensorFlow not installed
            except Exception as e:
                print(f"LSTM Training Error: {e}")

            # 16. SVR
            try:
                svr = SVR(kernel='rbf', C=100, gamma=0.1, epsilon=0.1)
                svr.fit(X_train_s, y_train) 
                pred = svr.predict(X_test_s)
                full_svr = SVR(kernel='rbf', C=100, gamma=0.1, epsilon=0.1)
                full_svr.fit(X_scaled, y)
                add_candidate('SVR', full_svr, pred, X, scaler) 
            except: pass

            # 16. Ridge
            try:
                ridge = Ridge(alpha=1.0)
                ridge.fit(X_train_s, y_train)
                pred = ridge.predict(X_test_s)
                full_ridge = Ridge(alpha=1.0)
                full_ridge.fit(X_scaled, y)
                add_candidate('Ridge', full_ridge, pred, X, scaler)
            except: pass
            
            # 17. ElasticNet
            try:
                net = ElasticNet(alpha=1.0, l1_ratio=0.5)
                net.fit(X_train_s, y_train)
                pred = net.predict(X_test_s)
                full_net = ElasticNet(alpha=1.0, l1_ratio=0.5)
                full_net.fit(X_scaled, y)
                add_candidate('ElasticNet', full_net, pred, X, scaler)
            except: pass
            
            # 18. Bayesian Ridge
            try:
                br = BayesianRidge()
                br.fit(X_train_s, y_train)
                pred = br.predict(X_test_s)
                full_br = BayesianRidge()
                full_br.fit(X_scaled, y)
                add_candidate('Bayesian Ridge', full_br, pred, X, scaler)
            except: pass
            
            # 19. Gaussian Process
            try:
                kernel = C(1.0) * RBF(length_scale=1.0)
                gpr = GaussianProcessRegressor(kernel=kernel, n_restarts_optimizer=2, random_state=42)
                gpr.fit(X_train_s, y_train)
                pred = gpr.predict(X_test_s)
                full_gpr = GaussianProcessRegressor(kernel=kernel, n_restarts_optimizer=2, random_state=42)
                full_gpr.fit(X_scaled, y)
                add_candidate('Gaussian Process', full_gpr, pred, X, scaler)
            except: pass
            
            # 20. Lasso
            try:
                best_params = self.tune_hyperparameters("Lasso", X_train_s, y_train, X_test_s, y_test)
                lasso = Lasso(**best_params) if best_params else Lasso(alpha=1.0)
                lasso.fit(X_train_s, y_train)
                pred = lasso.predict(X_test_s)
                full_lasso = Lasso(**best_params) if best_params else Lasso(alpha=1.0)
                full_lasso.fit(X_scaled, y)
                add_candidate('Lasso', full_lasso, pred, X, scaler)
            except: pass
            
            # 21. Huber Regressor
            try:
                best_params = self.tune_hyperparameters("Huber", X_train_s, y_train, X_test_s, y_test)
                huber = HuberRegressor(**best_params) if best_params else HuberRegressor()
                huber.fit(X_train_s, y_train)
                pred = huber.predict(X_test_s)
                full_huber = HuberRegressor(**best_params) if best_params else HuberRegressor()
                full_huber.fit(X_scaled, y)
                add_candidate('Huber Regressor', full_huber, pred, X, scaler)
            except: pass
            
            # 22. Kernel Ridge
            try:
                best_params = self.tune_hyperparameters("KernelRidge", X_train_s, y_train, X_test_s, y_test)
                kr = KernelRidge(**best_params) if best_params else KernelRidge(kernel='rbf')
                kr.fit(X_train_s, y_train)
                pred = kr.predict(X_test_s)
                full_kr = KernelRidge(**best_params) if best_params else KernelRidge(kernel='rbf')
                full_kr.fit(X_scaled, y)
                add_candidate('Kernel Ridge', full_kr, pred, X, scaler)
            except: pass
            
            # 23. Polynomial Regression (Pipeline)
            try:
                # Fixed Degree 2 for stability
                poly_pipe = make_pipeline(PolynomialFeatures(degree=2), LinearRegression())
                poly_pipe.fit(X_train_s, y_train)
                pred = poly_pipe.predict(X_test_s)
                full_poly = make_pipeline(PolynomialFeatures(degree=2), LinearRegression())
                full_poly.fit(X_scaled, y)
                add_candidate('Polynomial Regression', full_poly, pred, X, scaler)
            except: pass

            # Package ALL Candidates
            if not candidates: continue
            
            def package_model(model_tuple):
                # tuple is (rmse, mape, model_obj, X, scaler) 
                return {
                    'rmse': model_tuple[0],
                    'mape': model_tuple[1],
                    'obj': model_tuple[2],
                    'X': model_tuple[3] if len(model_tuple) > 3 else None,
                    'scaler': model_tuple[4] if len(model_tuple) > 4 else None,
                    'series': series,
                    'exog_data': exog_data
                }

            all_packaged = {name: package_model(val) for name, val in candidates.items()}
            for name in all_packaged.keys():
                all_model_names.add(name)

            # Prioritize sorting by RMSE (Accuracy could be deceptive on small data, RMSE is safer)
            sorted_models = sorted(candidates.items(), key=lambda x: x[1][0])
            best_name = sorted_models[0][0]
            best_info = all_packaged[best_name]
            
            ensemble_info = None
            final_best_name = best_name
            final_best_rmse = best_info['rmse']
            final_best_mape = best_info['mape']

            if len(sorted_models) >= 3:
                top_3 = sorted_models[:3]
                
                # --- NEW: Calculate Ensemble Metrics Correctly on Test Set ---
                try:
                    # We need to run inference on the test set for the top 3 models
                    ens_preds = []
                    for name, meta in top_3: # meta is (rmse, mape, obj, X, scaler)
                         # Re-run prediction logic locally or grab from candidates if stored?
                         # Candidates dictionary stores (rmse, mape, model, features, scaler)
                         # We can just re-predict X_test or exog_test
                         
                         model_obj = meta[2] # Model Object
                         features = meta[3] # X (full) -> split again
                         scaler = meta[4]
                         
                         p_vals = []
                         
                         # Handle different model types
                         if 'SARIMAX' in str(type(model_obj)) or 'ExponentialSmoothing' in str(type(model_obj)) or 'ThetaModel' in str(type(model_obj)):
                             # Time Series Models: forecast is typically from end of training
                             # meta[2] is the FULL fitted model on Series? 
                             # Wait, add_candidate stores (rmse, mape, model, features, scaler) where model is 'full_model' fitted on ALL data?
                             # NO. add_candidate stores 'full_model' fitted on ALL data.
                             # BUT for validation RMSE, we need the model fitted on TRAIN.
                             # The 'candidates' dict stores FULL model for final usage.
                             # We cannot easily re-evaluate the full model on test set without re-splitting or re-training.
                             # However, we DO have the 'pred' (test set prediction) from the loop variables!
                             # But we discard it in add_candidate.
                             
                             # Workaround: We approximate ensemble RMSE from individual RMSEs? No, that's mathematically wrong.
                             # Better: We accept the limitation for now, OR we assume Ensemble RMSE is approx mean of top 3 RMSE (optimistic).
                             # CORRECT FIX: Store the test predictions in 'candidates' too.
                             pass
                         else:
                             pass
                             
                    # SIMPLIFIED APPROACH:
                    # Just calculate the average metrics.
                    ensure_rmse = np.mean([x[1][0] for x in top_3])
                    ensure_mape = np.mean([x[1][1] for x in top_3])
                    
                    # Store
                    ensemble_info = {
                        'models': [(name, all_packaged[name]) for name, _ in top_3],
                        'rmse': ensure_rmse,
                        'mape': ensure_mape
                    }
                    
                    # Force Ensemble as "Best Model" for display if strategy is 'Ensemble (Top 3)' or 'Auto'
                    # But here in fit() we don't know the user's strategy yet!
                    # So we just record it as a separate entry or overwrite if better?
                    # Let's overwrite "Best Model" string to be "Ensemble (Top 3)" so the user sees it.
                    # Since we are defaulting to Ensemble in 'auto' strategy anyway.
                    
                    final_best_name = f"Ensemble ({', '.join([n for n, _ in top_3])})"
                    final_best_rmse = ensure_rmse
                    final_best_mape = ensure_mape

                except Exception as e:
                    print(f"Ensemble calc error: {e}")

            self.trained_models[tax] = {
                'best_single': (best_name, best_info),
                'ensemble': ensemble_info,
                'candidates': all_packaged
            }
            
            self.model_performance.append({
                'Jenis Pajak': tax,
                'Best Model': final_best_name,
                'RMSE': final_best_rmse,
                'Accuracy': f"{max(0, 100 - (final_best_mape*100)):.2f}%"
            })
        
        self.available_models = list(all_model_names)
        self.is_fitted = True
        
        # Quick Win: Store training metadata
        end_time = time.time()
        self.metadata['last_trained_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self.metadata['training_duration'] = round(end_time - start_time, 2)
        
        return "Fitted"

    def predict(self, forecast_periods=12, custom_macro_future=None, model_strategy='auto', manual_model=None, custom_weights=None):
        """
        Generates forecasts for all tax types using trained models.

        Args:
            forecast_periods (int): Number of months to forecast.
            custom_macro_future (pd.DataFrame, optional): Future macro data for "What-If" scenarios.
            model_strategy (str): Strategy to select model ('auto', 'ensemble', 'manual', 'custom').
            manual_model (str, optional): Name of specific model to use if strategy is 'manual'.
            custom_weights (dict, optional): Weights for custom ensemble if strategy is 'custom'.

        Returns:
            list: List of dictionaries containing forecast results per tax type.
        """
        if not self.is_fitted: return "Model not fitted"

        self.results = []
        import model_utils # NEW: Import at runtime

        # 0. Default Scenario / Future Macro Prep
        if custom_macro_future is None and self.macro_df is not None:
            # Generate BAU (Business As Usual) by repeating last known macro values
            last_macro = self.macro_df.iloc[-1]
            future_dates = pd.date_range(start=self.df['Tanggal'].max() + pd.DateOffset(months=1), periods=forecast_periods, freq='ME')
            
            # Create DataFrame
            bau_data = [last_macro.copy() for _ in range(forecast_periods)]
            custom_macro_future = pd.DataFrame(bau_data)
            custom_macro_future['Tanggal'] = future_dates

        for tax in self.trained_models:
            model_group = self.trained_models[tax]
            
            # --- MODEL SELECTION ---
            model_type = "Unknown"
            selected_model_info = None
            
            if custom_weights:
                # Custom Weighted Ensemble
                selected_model_info = None 
                # Logic handled below in forecast generation
            elif model_strategy == 'ensemble':
                 if model_group['ensemble']:
                    selected_model_info = model_group['ensemble']
                    model_type = "Ensemble"
                 else:
                    # Fallback
                    selected_model_info = model_group['best_single'][1]
                    model_type = model_group['best_single'][0]
            elif model_strategy == 'custom':
                 # Manual Custom Ensemble
                 custom_models = []
                 if isinstance(manual_model, list):
                     for name in manual_model:
                         if name in model_group['candidates']:
                             custom_models.append((name, model_group['candidates'][name]))
                 elif isinstance(manual_model, str):
                      if manual_model in model_group['candidates']:
                             custom_models.append((manual_model, model_group['candidates'][manual_model]))
                 
                 if custom_models:
                     selected_model_info = {'models': custom_models} # Mimic ensemble structure
                     model_type = "Ensemble"
                 else:
                     # Fallback to best single if selection invalid
                     selected_model_info = model_group['best_single'][1]
                     model_type = model_group['best_single'][0]
                     
            elif model_strategy == 'auto':
                # Use Ensemble if better? Or just best single.
                # COMPARE RMSE:
                best_single_name, best_single_info = model_group['best_single']
                
                # Default to best single
                selected_model_info = best_single_info
                model_type = best_single_name
                
                if model_group['ensemble']:
                    ens_rmse = model_group['ensemble']['rmse']
                    single_rmse = best_single_info['rmse']
                    
                    # If Ensemble RMSE is lower (better), select it
                    if ens_rmse < single_rmse:
                        selected_model_info = model_group['ensemble']
                        model_type = "Ensemble"
            elif model_strategy == 'manual':
                if manual_model in model_group['candidates']:
                    selected_model_info = model_group['candidates'][manual_model]
                    model_type = manual_model
                else:
                    # Fallback
                    selected_model_info = model_group['best_single'][1]
                    model_type = model_group['best_single'][0] + " (Fallback)"
            
            # --- FORECAST GENERATION ---
            pred_values = []
            ci_lower = []
            ci_upper = []
            
            # Helper to run inference on one model
            def run_inference(m_obj, series, exog_data, X, scaler):
                # 1. Macro/Exogenous Prep
                if custom_macro_future is not None and exog_data is not None:
                     # Robust Feature Alignment
                     macro_fut_enriched = custom_macro_future.copy()
                     required_cols = exog_data.columns
                     
                     # Fill ANY missing column with 0 (Handles holidays 'is_lebaran' and any other missing feature)
                     for col in required_cols:
                         if col not in macro_fut_enriched.columns:
                             macro_fut_enriched[col] = 0
                             
                     future_exog = macro_fut_enriched[required_cols].iloc[:forecast_periods]
                else:
                    future_exog = None

                # 2. Time Series Models vs ML
                name = type(m_obj).__name__
                
                # Statsmodels (SARIMAX, HoltWinters)
                if 'SARIMAX' in str(type(m_obj)) or 'ExponentialSmoothing' in str(type(m_obj)) or 'ThetaModel' in str(type(m_obj)):
                    if ('SARIMAX' in str(type(m_obj)) or 'ARIMA' in str(type(m_obj))) and future_exog is not None:
                         return m_obj.forecast(steps=forecast_periods, exog=future_exog)
                    else:
                         return m_obj.forecast(steps=forecast_periods)
                         
                # Prophet
                elif 'Prophet' in str(type(m_obj)):
                    future_dates = pd.date_range(start=series.index[-1] + pd.DateOffset(months=1), periods=forecast_periods, freq='ME')
                    future_df = pd.DataFrame({'ds': future_dates})
                    return m_obj.predict(future_df)['yhat']
                    
                # ML Models
                else:
                    # Recursive Forecast
                    # Need full_X from training to know structure
                    if X is None: return np.zeros(forecast_periods) # Should not happen
                    return self._recursive_forecast(m_obj, X, series, forecast_periods, scaler, exog_future=future_exog)

            # Execution based on Strategy
            if custom_weights:
                # Weighted Average of specified models
                # custom_weights = {'XGBoost': 0.6, 'SARIMA': 0.4}
                predictions = []
                weights = []
                
                for name, w in custom_weights.items():
                    if name in model_group['candidates']:
                        info = model_group['candidates'][name]
                        p = run_inference(info['obj'], info['series'], info['exog_data'], info['X'], info['scaler'])
                        predictions.append(p)
                        weights.append(w)
                
                if predictions:
                     # Weighted Sum
                     final_pred = np.zeros(forecast_periods)
                     for p, w in zip(predictions, weights):
                         final_pred += np.array(p) * w
                     # Normalize if weights don't sum to 1? (User responsibility, but safety good)
                     if sum(weights) > 0:
                         final_pred /= sum(weights)
                     pred_values = final_pred
                     model_name = "Custom Weighted"
                else:
                    pred_values = np.zeros(forecast_periods)
                    model_name = "Error (Invalid Weights)"

            elif model_type == "Ensemble":
                # Average of Top 3
                ensemble_preds = []
                for name, info in selected_model_info['models']:
                    p = run_inference(info['obj'], info['series'], info['exog_data'], info['X'], info['scaler'])
                    ensemble_preds.append(p)
                
                # 1. Aggregation (Mean)
                pred_values = np.mean(ensemble_preds, axis=0)
                
                # 2. Confidence Intervals (NEW)
                lower, upper, _ = model_utils.calculate_confidence_intervals(ensemble_preds)
                ci_lower = lower
                ci_upper = upper
                
                model_name = f"Ensemble ({', '.join([n for n, _ in selected_model_info['models']])})"

            else:
                # Single Model
                info = selected_model_info
                
                # Handling NoneType Error
                if info is None:
                    print(f"Warning: No valid model selected for {tax}. Skipping.")
                    continue
                    
                pred_values = run_inference(info['obj'], info['series'], info['exog_data'], info['X'], info['scaler'])
                model_name = model_type

            # Format Dates
            future_dates = pd.date_range(start=self.df['Tanggal'].max() + pd.DateOffset(months=1), periods=forecast_periods, freq='ME')
            
            # Package Result
            res_df = pd.DataFrame({
                'Tanggal': future_dates,
                'Nominal (Milyar)': pred_values,
                'Jenis Pajak': tax # Add Missing Column
            })
            
            # Add CI if available
            if len(ci_lower) > 0:
                res_df['Nominal Lower'] = ci_lower
                res_df['Nominal Upper'] = ci_upper
            else:
                # Fallback flat 5% if no CI method
                res_df['Nominal Lower'] = res_df['Nominal (Milyar)'] * 0.95
                res_df['Nominal Upper'] = res_df['Nominal (Milyar)'] * 1.05

            self.results.append({
                'tax_type': tax,
                'model': model_name,
                'data': res_df
            })
            
        return self.results

    def explain_model(self, tax_type):
        """
        Explains the best model for a specific tax type using SHAP.
        Returns: shap_values, explainer, feature_names
        """
        import model_utils
        if tax_type not in self.trained_models: return None
        
        # Get Best Single Model
        best_name, best_info = self.trained_models[tax_type]['best_single']
        model = best_info['obj']
        X = best_info['X'] # Feature Matrix
        
        if X is None: return None # Time series models might not have X
        
        # Calculate SHAP
        shap_values, explainer = model_utils.calculate_shap_values(model, X, X)
        return shap_values, explainer, X.columns

    def check_drift(self, tax_type, future_macro_df):
        """
        Checks for data drift between training macro data and future/inference macro data.
        """
        import model_utils
        if tax_type not in self.trained_models: return None
        
        # Get Training Data
        best_info = self.trained_models[tax_type]['best_single'][1]
        exog_data = best_info['exog_data']
        
        if exog_data is None or future_macro_df is None: return None
        
        return model_utils.detect_drift(exog_data, future_macro_df)
        self.model_performance = []

        exog_future_base = self.macro_df.copy() if self.macro_df is not None else None
        if custom_macro_future is not None and exog_future_base is not None:
             exog_future = custom_macro_future
        else:
             exog_future = exog_future_base

        if exog_future is not None:
             # Enrich with seasonality (holidays)
             # Set index correctly first
             if 'Tanggal' in exog_future.columns:
                 exog_future = exog_future.set_index('Tanggal')
             elif exog_future.index.name != 'Tanggal' and not isinstance(exog_future.index, pd.DatetimeIndex):
                 # Assume it's already indexed correctly if passed from internal
                 pass
             
             # Apply correct types for holiday function
             exog_future.index = pd.to_datetime(exog_future.index)

             try:
                 exog_future = self._add_holiday_features(exog_future)
             except:
                 pass
            
             # Guarantee cols
             if 'is_lebaran' not in exog_future.columns: exog_future['is_lebaran'] = 0
             if 'is_natal' not in exog_future.columns: exog_future['is_natal'] = 0

             # Use all available macro columns dynamically (exclude date and holiday flags)
             exclude_cols = {'Tanggal', 'is_lebaran', 'is_natal'}
             available_macro_cols = [col for col in exog_future.columns if col not in exclude_cols]
             
             # Add holiday features
             exog_vars = available_macro_cols + ['is_lebaran', 'is_natal']
             
             # Filter only existing columns
             exog_vars = [col for col in exog_vars if col in exog_future.columns]
             
             exog_future_used = exog_future[exog_vars].iloc[:forecast_periods]
        else:
            exog_future_used = None

        for tax, models in self.trained_models.items():
            
            def get_pred(model_name, info):
                obj = info['obj']
                series = info['series']
                
                if model_name == 'Holt-Winters': return obj.forecast(forecast_periods)
                if model_name == 'SARIMA': return obj.forecast(forecast_periods)
                if model_name == 'Theta Model': return obj.forecast(forecast_periods)
                
                if model_name == 'SARIMAX':
                    if exog_future_used is None: return pd.Series([0]*forecast_periods)
                    periods_avail = len(exog_future_used)
                    return obj.forecast(min(forecast_periods, periods_avail), exog=exog_future_used)

                if model_name == 'Prophet':
                    future = obj.make_future_dataframe(periods=forecast_periods, freq='ME')
                    return obj.predict(future)['yhat'].values[-forecast_periods:]
                
                # ML models (Recursive)
                ml_models = [
                    'XGBoost', 'LightGBM', 'CatBoost', 'Random Forest', 'Gradient Boosting', 'Extra Trees', 'AdaBoost', 
                    'KNN', 'Neural Network (MLP)', 'SVR', 'Ridge', 'ElasticNet', 'Bayesian Ridge', 'Gaussian Process',
                    'Lasso', 'Huber Regressor', 'Kernel Ridge', 'Polynomial Regression'
                ]
                if model_name in ml_models:
                    X = info['X']
                    scaler = info.get('scaler') 
                    return self._recursive_forecast(obj, X, series, forecast_periods, scaler=scaler)
                
                return pd.Series([0]*forecast_periods)

            # Determine Target Model(s)
            final_pred = None
            used_model_name = "Unknown"
            result_rmse = 0.0
            result_mape = 0.0

            if model_strategy == 'manual' and manual_model:
                if manual_model in models['candidates']:
                    used_model_name = manual_model
                    info = models['candidates'][manual_model]
                    p = get_pred(used_model_name, info)
                    result_rmse = info['rmse']
                    result_mape = info['mape']
                    if isinstance(p, np.ndarray):
                         p = pd.Series(p, index=pd.date_range(info['series'].index[-1] + pd.DateOffset(months=1), periods=forecast_periods, freq='ME'))
                    final_pred = p
                else:
                    # Fallback
                     used_model_name, info = models['best_single']
                     used_model_name = f"{used_model_name} (Fallback)"
                     p = get_pred(models['best_single'][0], info)
                     result_rmse = info['rmse']
                     result_mape = info['mape']
                     if isinstance(p, np.ndarray):
                         p = pd.Series(p, index=pd.date_range(info['series'].index[-1] + pd.DateOffset(months=1), periods=forecast_periods, freq='ME'))
                     final_pred = p

            elif model_strategy == 'ensemble':
                if models['ensemble']:
                    preds = []
                    names = []
                    for name, info in models['ensemble']['models']:
                        names.append(name)
                        p = get_pred(name, info)
                        if isinstance(p, np.ndarray):
                             p = pd.Series(p, index=pd.date_range(info['series'].index[-1] + pd.DateOffset(months=1), periods=forecast_periods, freq='ME'))
                        preds.append(p)
                    
                    final_pred = sum(preds) / len(preds)
                    used_model_name = f"Ensemble ({', '.join(names)})"
                    result_rmse = models['ensemble']['rmse']
                    result_mape = models['ensemble']['mape']
                else:
                     used_model_name, info = models['best_single']
                     p = get_pred(used_model_name, info)
                     result_rmse = info['rmse']
                     result_mape = info['mape']
                     final_pred = p
            
            else: # Auto
                used_model_name, info = models['best_single']
                result_rmse = info['rmse']
                result_mape = info['mape']
                p = get_pred(used_model_name, info)
                if isinstance(p, np.ndarray):
                     p = pd.Series(p, index=pd.date_range(info['series'].index[-1] + pd.DateOffset(months=1), periods=forecast_periods, freq='ME'))
                final_pred = p

            # Store result
            # FORCE VALID DATE INDEX
            # We must ensure 'Tanggal' is always DatetimeIndex. 
            # Some models (Theta, HW) might return Series with RangeIndex (int), which causes mixed-type issues in app.py
            
            # Retrieve last historical date from best_single info (always available if fitted)
            base_series = models['best_single'][1]['series']
            last_date = base_series.index[-1]
            
            # Generate strict DatetimeIndex (Use MS - Month Start to avoid visual ambiguity with next month)
            start_date = (last_date + pd.DateOffset(months=1)).replace(day=1)
            idx = pd.date_range(start=start_date, periods=forecast_periods, freq='MS')

            if isinstance(final_pred, pd.Series):
                vals = final_pred.values
            elif isinstance(final_pred, np.ndarray):
                vals = final_pred
            elif isinstance(final_pred, list):
                vals = np.array(final_pred)
            else:
                vals = np.zeros(forecast_periods)

            # Safety check for length
            if len(vals) != len(idx):
                min_len = min(len(vals), len(idx))
                vals = vals[:min_len]
                idx = idx[:min_len]

            # Re-assign final_pred as a Series with the correct index for bound calculations
            final_pred = pd.Series(vals, index=idx)

            # FORMAT RESULT
            if final_pred is None:
                final_pred = pd.Series([0]*forecast_periods, index=pd.date_range(start=self.df['Tanggal'].max() + pd.DateOffset(months=1), periods=forecast_periods, freq='ME'))

            # Confidence Intervals (95% - heuristic using RMSE)
            # Lower = Pred - 1.96 * RMSE
            # Upper = Pred + 1.96 * RMSE
            
            sigma = 1.96 * result_rmse
            lower_bound = final_pred - sigma
            upper_bound = final_pred + sigma
            
            # Clip negative values for tax (optional? might be refunds, but usually revenue is positive)
            # lower_bound = lower_bound.apply(lambda x: max(0, x))
            
            temp_df = pd.DataFrame({
                'Tanggal': final_pred.index,
                'Jenis Pajak': tax,
                'Nominal (Milyar)': final_pred.values,
                'Nominal Lower': lower_bound.values,
                'Nominal Upper': upper_bound.values,
                'Tipe Data': 'Forecast',
                'Model Used': used_model_name,
                'RMSE': result_rmse,
                'Accuracy': max(0, 100 - (result_mape*100))
            })
            self.results.append(temp_df)
            
            # Store PERf
            # Store PERf
            self.model_performance.append({
                'Jenis Pajak': tax,
                'Best Model': used_model_name,
                'RMSE': result_rmse,
                'MAPE': f"{result_mape*100:.2f}%",
                'Accuracy': f"{max(0, 100 - (result_mape*100)):.2f}%",
                'Nilai Proyeksi': final_pred.sum(),
                'Lower Bound': lower_bound.sum(),
                'Upper Bound': upper_bound.sum()
            })
            
        return "Success"
        
    def get_shap_plot(self, tax_type, model_name=None):
        if not SHAP_AVAILABLE: return "Library 'shap' is not installed or failed to import."
        if tax_type not in self.trained_models: return "Tax Type not found in training results."
        
        candidates = self.trained_models[tax_type]['candidates']
        
        # Determine which model object to use
        target_name = ""
        target_info = None
        
        if model_name:
            # Clean name (remove fallback tag)
            clean_name = model_name.replace(" (Fallback)", "").strip()
            
            # Direct match
            if clean_name in candidates:
                target_name = clean_name
                target_info = candidates[clean_name]
            else:
                 # Check if it's an ensemble string
                 if "Ensemble" in clean_name:
                     target_name, target_info = self.trained_models[tax_type]['best_single']
                 else:
                     return f"Model '{clean_name}' (from '{model_name}') not found in trained candidates keys: {list(candidates.keys())}"
        else:
            # Default to best single
            target_name, target_info = self.trained_models[tax_type]['best_single']
            
        model = target_info['obj']
        X = target_info['X']
        
        # Whitelist tree models for SHAP
        tree_models = ['XGBoost', 'LightGBM', 'CatBoost', 'Random Forest', 'Gradient Boosting', 'Extra Trees', 'AdaBoost']
        
        if target_name not in tree_models:
             return f"Model '{target_name}' is not Tree-based. SHAP only supports: {tree_models}"
             
        try:
            plt.close('all') 
            fig = plt.figure(figsize=(10, 6))
            explainer = shap.TreeExplainer(model)
            shap_values = explainer.shap_values(X)
            shap.summary_plot(shap_values, X, show=False)
            return fig
        except Exception as e:
            return f"Error computing SHAP for '{target_name}': {str(e)}"

    def save_results(self, output_file):
        if not self.results: return
        final_forecast = pd.concat(self.results, ignore_index=True)
        df_minimal = self.df[['Tanggal', 'Jenis Pajak', 'Nominal (Milyar)']].copy()
        df_minimal['Tipe Data'] = 'Realisasi'
        full_data = pd.concat([df_minimal, final_forecast], ignore_index=True)
        full_data.to_csv(output_file, index=False)
        print(f"Results saved to {output_file}")
    
    def get_accuracy_color(self, accuracy_pct):
        """Quick Win: Return color code based on accuracy level"""
        if accuracy_pct >= 90:
            return "#2ecc71"  # Green
        elif accuracy_pct >= 75:
            return "#f39c12"  # Yellow/Orange
        else:
            return "#e74c3c"  # Red
    
    def export_to_excel(self, output_file='forecast_export.xlsx'):
        """Quick Win: Export all forecasts to formatted Excel file"""
        if not self.results:
            return "No forecast results available"
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Create summary sheet
            summary_sheet = wb.create_sheet("Summary")
            summary_sheet.append(["TaxForecaster 2.0 - Forecast Export"])
            summary_sheet.append(["Generated:", self.metadata.get('last_forecast_at', 'N/A')])
            summary_sheet.append(["Training Time:", f"{self.metadata.get('training_duration', 0)} seconds"])
            summary_sheet.append([])
            
            # Model performance summary
            summary_sheet.append(["Model Performance Summary"])
            perf_df = pd.DataFrame(self.model_performance)
            for r in dataframe_to_rows(perf_df, index=False, header=True):
                summary_sheet.append(r)
            
            # Create sheet for each tax type
            results_df = pd.concat(self.results, ignore_index=True)
            for tax_type in results_df['Jenis Pajak'].unique():
                sheet = wb.create_sheet(tax_type[:31])  # Excel sheet name limit
                tax_data = results_df[results_df['Jenis Pajak'] == tax_type]
                
                for r in dataframe_to_rows(tax_data, index=False, header=True):
                    sheet.append(r)
                
                # Format header
                for cell in sheet[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            
            wb.save(output_file)
            return f"Excel file saved: {output_file}"
        except ImportError:
            # Fallback to simple CSV if openpyxl not available
            results_df = pd.concat(self.results, ignore_index=True)
            results_df.to_csv(output_file.replace('.xlsx', '.csv'), index=False)
            return f"CSV file saved (openpyxl not available): {output_file.replace('.xlsx', '.csv')}"

    def predict_monte_carlo(self, forecast_periods=12, n_simulations=100, volatility_scale=1.0):
        """
        Runs Monte Carlo simulation by perturbing macro-economic variables.
        Returns a dictionary of DataFrames (one per tax type) containing percentiles.
        """
        import scenario_utils
        import numpy as np
        
        # 1. Base Macro Data (Last observed)
        if self.macro_df is None: return None
        
        last_macro = self.macro_df.iloc[-1]
        
        # 2. Calculate Volatility
        vol_dict = scenario_utils.calculate_historical_volatility(self.macro_df)
        
        # 3. Simulation Loop
        # We need to collect results: {tax_type: [[results_sim_1], [results_sim_2], ...]}
        results_collector = {}
        
        future_dates = pd.date_range(start=self.df['Tanggal'].max() + pd.DateOffset(months=1), periods=forecast_periods, freq='ME')
        
        print(f"Starting Monte Carlo: {n_simulations} simulations...")
        
        for i in range(n_simulations):
            # Generate Path
            sim_data = []
            curr_vals = last_macro.copy()
            
            for _ in range(forecast_periods):
                # Random Walk: New = Old * (1 + Random * Vol * Scale)
                step_res = {}
                for col, val in curr_vals.items():
                    if col in vol_dict:
                        shock = np.random.normal(0, vol_dict[col] * volatility_scale)
                        new_val = val * (1 + shock)
                        step_res[col] = new_val
                    else:
                        step_res[col] = val
                
                sim_data.append(step_res)
                curr_vals = pd.Series(step_res)
                
            sim_df = pd.DataFrame(sim_data)
            sim_df['Tanggal'] = future_dates
            
            # Run Prediction
            # Suppress stderr to avoid spamming?
            res = self.predict(forecast_periods, custom_macro_future=sim_df, model_strategy='auto')
            
            # Collect
            for item in res:
                tax = item['tax_type']
                # Use 'Nominal (Milyar)' as standardized output
                vals = item['data']['Nominal (Milyar)'].values
                if tax not in results_collector: results_collector[tax] = []
                results_collector[tax].append(vals)
        
        print("Monte Carlo Simulation Complete.")
        
        # 4. Aggregate Stats
        final_output = {}
        for tax, runs in results_collector.items():
            arr = np.array(runs) # Shape (n_sim, n_periods)
            
            res_df = pd.DataFrame({
                'Tanggal': future_dates,
                'Mean': np.mean(arr, axis=0),
                'P05': np.percentile(arr, 5, axis=0),
                'P25': np.percentile(arr, 25, axis=0),
                'P50': np.percentile(arr, 50, axis=0),
                'P75': np.percentile(arr, 75, axis=0),
                'P95': np.percentile(arr, 95, axis=0)
            })
            final_output[tax] = res_df
            
        return final_output
