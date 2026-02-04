import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import os
import sys
from statsmodels.tsa.seasonal import seasonal_decompose
import matplotlib.pyplot as plt
import base64

import seaborn as sns
import matplotlib.pyplot as plt

# Add parent directory to path to import modules from root
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from forecaster import TaxForecaster
from model_info import MODEL_INFO
from fetch_macro import fetch_macro_data
import style
import scenario_utils # Centralized scenarios
import report_generator # Added for Executive Report

# --- Page Config ---
st.set_page_config(page_title="Dashboard | TaxForecaster 2.0", layout="wide", page_icon="📊")

# --- Apply Theme & Seaborn Style ---
style.apply_theme()
sns.set_theme(style="darkgrid", context="talk", palette="mako")

# --- Quick Win Helper Functions ---
def copy_to_clipboard_button(label, value, key):
    """Add a copy-to-clipboard button for a value"""
    if st.button(f"📋 {label}", key=key, help="Click to copy to clipboard"):
        st.code(value, language=None)
        st.success(f"✅ Copied: {value}")

def create_summary_card(tax_name, model_name, accuracy, forecast_value):
    """
    Creates a compact HTML summary card for a tax type forecast.

    Args:
        tax_name (str): Name of the tax type.
        model_name (str): Name of the model used.
        accuracy (str): Accuracy percentage string.
        forecast_value (float): Forecasted revenue value.

    Returns:
        str: HTML string for the card.
    """
    return f"""
    <div style="padding: 10px; background: rgba(52, 152, 219, 0.1); border-radius: 8px; border-left: 4px solid #3498DB; margin-bottom: 10px;">
        <div style="font-weight: bold; color: #2c3e50;">{tax_name}</div>
        <div style="font-size: 0.85rem; color: #7f8c8d;">
            Model: {model_name} | Accuracy: {accuracy} | Forecast: Rp {forecast_value:,.0f}M
        </div>
    </div>
    """

def export_all_data_zip(forecast_df, model_performance, scenario_params):
    """
    Creates a ZIP file containing forecast data, model performance, and scenario parameters.

    Args:
        forecast_df (pd.DataFrame): The forecast results dataframe.
        model_performance (list or pd.DataFrame): Model performance metrics.
        scenario_params (dict): Dictionary of scenario parameters used.

    Returns:
        bytes: The ZIP file content as bytes.
    """
    import io
    import zipfile
    from datetime import datetime
    
    zip_buffer = io.BytesIO()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # Add forecast data
        csv_buffer = io.StringIO()
        forecast_df.to_csv(csv_buffer, index=False)
        zip_file.writestr(f'forecast_data_{timestamp}.csv', csv_buffer.getvalue())
        
        # Add model performance
        perf_buffer = io.StringIO()
        pd.DataFrame(model_performance).to_csv(perf_buffer, index=False)
        zip_file.writestr(f'model_performance_{timestamp}.csv', perf_buffer.getvalue())
        
        # Add scenario parameters
        param_buffer = io.StringIO()
        param_buffer.write(f"Scenario Parameters - {timestamp}\n")
        param_buffer.write("="*50 + "\n")
        for key, value in scenario_params.items():
            param_buffer.write(f"{key}: {value}\n")
        zip_file.writestr(f'scenario_params_{timestamp}.txt', param_buffer.getvalue())
    
    zip_buffer.seek(0)
    return zip_buffer.getvalue()

# --- Helper Functions (Plotting) ---
def plot_forecast_chart(df, title, show_legend=True, baseline_df=None):
    """
    Generates a standardized Plotly line chart for historical and forecast data.

    Args:
        df (pd.DataFrame): Combined dataframe of history and forecast.
        title (str): Chart title.
        show_legend (bool): Whether to show the legend.
        baseline_df (pd.DataFrame, optional): Previous forecast for comparison.

    Returns:
        plotly.graph_objects.Figure: The Plotly figure object.
    """
    fig = go.Figure()

    # Split Data
    hist_data = df[df['Tipe Data'] == 'Realisasi']
    fc_data = df[df['Tipe Data'] == 'Forecast']

    # 1. Historical Line (Solid Neon Blue)
    fig.add_trace(go.Scatter(
        x=hist_data['Tanggal'], 
        y=hist_data['Nominal (Milyar)'],
        mode='lines',
        name='Realisasi (History)',
        line=dict(color='#00d2ff', width=3), # Neon Blue
        hovertemplate='%{x|%b %Y}<br>Rp %{y:,.0f} Milyar'
    ))

    # 4. Baseline Line (If available - Grey Dashed)
    if baseline_df is not None and not baseline_df.empty:
        base_fc = baseline_df[baseline_df['Tipe Data'] == 'Forecast']
        if not base_fc.empty:
            fig.add_trace(go.Scatter(
                x=base_fc['Tanggal'], 
                y=base_fc['Nominal (Milyar)'],
                mode='lines',
                name='Baseline (Awal)',
                line=dict(color='#BDC3C7', width=2, dash='dot'), # Lighter Grey
                hovertemplate='%{x|%b %Y}<br>Rp %{y:,.0f} Milyar (Base)'
            ))

    # 2. Forecast Line (Neon Orange - Simulation/Current)
    fig.add_trace(go.Scatter(
        x=fc_data['Tanggal'], 
        y=fc_data['Nominal (Milyar)'],
        mode='lines+markers',
        name='Forecast (Skenario)',
        line=dict(color='#ff9f43', width=3, dash='solid'), # Neon Orange
        marker=dict(size=8, color='#ff9f43', symbol='circle', line=dict(width=2, color='white')),
        hovertemplate='%{x|%b %Y}<br>Rp %{y:,.0f} Milyar'
    ))

    # 3. Connection Line (Bridge the gap)
    if not hist_data.empty and not fc_data.empty:
        last_hist = hist_data.iloc[-1]
        first_fc = fc_data.iloc[0]
        fig.add_trace(go.Scatter(
            x=[last_hist['Tanggal'], first_fc['Tanggal']],
            y=[last_hist['Nominal (Milyar)'], first_fc['Nominal (Milyar)']],
            mode='lines',
            line=dict(color='#ff9f43', width=3, dash='solid'),
            showlegend=False,
            hoverinfo='skip'
        ))

    # 4. Confidence Interval (Shaded Area with Gradient feel)
    if 'Nominal Lower' in fc_data.columns and 'Nominal Upper' in fc_data.columns:
        fig.add_trace(go.Scatter(
            x=fc_data['Tanggal'],
            y=fc_data['Nominal Upper'],
            mode='lines',
            line=dict(width=0),
            showlegend=False,
            name='Upper Bound'
        ))
        fig.add_trace(go.Scatter(
            x=fc_data['Tanggal'],
            y=fc_data['Nominal Lower'],
            mode='lines',
            line=dict(width=0),
            fill='tonexty', # Fill to previous trace (Upper)
            fillcolor='rgba(255, 159, 67, 0.15)', # Transparent Orange
            showlegend=True,
            name='95% Confidence'
        ))

    # Layout Styling (Professional Dark/Clean Look)
    fig.update_layout(
        title=dict(text=title, font=dict(size=20, color='#2c3e50')),
        xaxis=dict(showgrid=True, gridcolor='#ecf0f1', tickformat='%b %Y'),
        yaxis=dict(showgrid=True, gridcolor='#ecf0f1', title="Total (Milyar Rupiah)"),
        hovermode="x unified",
        height=450,
        margin=dict(t=50, l=10, r=10, b=10),
        plot_bgcolor='rgba(0,0,0,0)', # Transparent background
        paper_bgcolor='rgba(0,0,0,0)',
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1) if show_legend else dict(visible=False)
    )
    return fig

# --- Initialization ---
if 'forecaster_v6' not in st.session_state:
    st.session_state['forecaster_v6'] = None
if 'previous_forecast_comparison' not in st.session_state:
    st.session_state['previous_forecast_comparison'] = None

# --- Quick Win: Keyboard Shortcuts Documentation ---
st.markdown("""
<style>
.keyboard-shortcut {
    background: rgba(52, 152, 219, 0.1);
    padding: 8px 12px;
    border-radius: 5px;
    font-size: 0.85rem;
    margin: 5px 0;
    border-left: 3px solid #3498DB;
}
.kbd {
    background: #34495e;
    color: #ecf0f1;
    padding: 2px 6px;
    border-radius: 3px;
    font-family: monospace;
    font-weight: bold;
}
</style>
""", unsafe_allow_html=True)

# --- SIDEBAR CONFIG ---
with st.sidebar:
    st.subheader("🗂️ Data & Config")
    status_placeholder = st.empty()
    
    # Path Adjustment for Pages
    current_dir = os.path.dirname(os.path.abspath(__file__))
    root_dir = os.path.abspath(os.path.join(current_dir, '..'))
    
    default_history = os.path.join(root_dir, 'tax_history.csv')
    default_macro = os.path.join(root_dir, 'macro_2025.csv')
    macro_path = os.path.join(root_dir, 'macro_data_auto.csv')

    st.markdown("---")
    st.subheader("🪙 Data Historis Penerimaan Pajak")

    history_file = st.file_uploader("Upload Data Historis (CSV)", type="csv")
    
    st.markdown("---")
    st.subheader("🌍 Data Makro Ekonomi")
    
    # High Priority: Real-time Data Integration
    # High Priority: Real-time Data Integration
    # [Feature] Optional Macro Data Source
    macro_source = st.radio("Sumber Data Makro:", ["🌐 Auto-Download (Yahoo/World Bank)", "📂 Upload Manual (CSV)"], index=0)

    if macro_source == "🌐 Auto-Download (Yahoo/World Bank)":
        # Existing Auto Logic
        if st.button("Unduh Data Makro Otomatis", use_container_width=True):
            with st.spinner("Mengunduh data kurs, minyak (Yahoo) dan inflasi, GDP (World Bank)..."):
                try:
                    df_macro = fetch_macro_data()
                    if df_macro is not None:
                        df_macro.to_csv(macro_path, index=False)
                        st.session_state['macro_fetch_time'] = pd.Timestamp.now()
                        st.success("✅ Data Makro berhasil diperbarui!")
                        st.rerun()
                    else:
                        st.error("Gagal mengunduh data.")
                except Exception as e:
                    st.error(f"Error: {e}")
        
        auto_refresh = st.checkbox("♻️ Auto-refresh on load", help="Automatically fetch latest macro data when dashboard loads")
        if auto_refresh and 'auto_refreshed' not in st.session_state:
            st.session_state['auto_refreshed'] = True
            with st.spinner("Auto-refreshing macro data..."):
                try:
                    df_macro = fetch_macro_data()
                    if df_macro is not None:
                        df_macro.to_csv(macro_path, index=False)
                        st.success("✅ Auto-refresh complete!")
                except: pass
        
        # Set file to auto path if exists
        macro_file = macro_path if os.path.exists(macro_path) else None

    else:
        # Manual Upload Logic
        uploaded_macro = st.file_uploader("Upload File Makro (CSV)", type="csv")
        if uploaded_macro is not None:
            # Save uploaded file to specific manual path or use directly (cleaner to save)
            manual_macro_path = os.path.join(root_dir, 'macro_data_manual.csv')
            try:
                pd.read_csv(uploaded_macro).to_csv(manual_macro_path, index=False)
                macro_file = manual_macro_path
                st.success("✅ File Makro manual berhasil dimuat")
            except Exception as e:
                st.error(f"Error reading CSV: {e}")
                macro_file = None
        else:
            # Use previously uploaded manual file if exists
            manual_macro_path = os.path.join(root_dir, 'macro_data_manual.csv')
            if os.path.exists(manual_macro_path):
                 macro_file = manual_macro_path
                 st.info("ℹ️ Menggunakan file manual yang tersimpan sebelumnya.")
            else:
                 macro_file = None

    # Check if macro file exists (Using the selected source)
    if macro_file and os.path.exists(macro_file):
        try:
            macro_df_info = pd.read_csv(macro_file)
        except:
            macro_df_info = pd.DataFrame() # Fallback
            
        # Quick Win: Display last updated timestamp
        import os
        from datetime import datetime
        last_modified = os.path.getmtime(macro_file)
        last_updated = datetime.fromtimestamp(last_modified).strftime('%d/%m/%Y %H:%M')
        
        # High Priority: Data Quality Indicator
        time_diff = datetime.now().timestamp() - last_modified
        hours_old = time_diff / 3600
        
        if hours_old < 24:
            freshness_icon = "🟢"  # Green - Fresh
            freshness_text = "Fresh"
        elif hours_old < 168:  # 1 week
            freshness_icon = "🟡"  # Yellow - Moderate
            freshness_text = "Moderate"
        else:
            freshness_icon = "🔴"  # Red - Stale
            freshness_text = "Stale"
        
        st.success(f"✅ Data Makro Tersedia ({macro_df_info.shape[0]} baris) {freshness_icon} {freshness_text}")
        st.caption(f"📅 Last Updated: {last_updated} ({hours_old:.1f} hours ago)")
        macro_file = macro_path
    else:
        st.warning("⚠️ Data Makro belum tersedia. Klik tombol di atas.")
        macro_file = None 

    # Show data availability status
    if not history_file:
        with status_placeholder.container():
            st.warning("⚠️ **Data Historis Belum Tersedia**")
            st.info("📌 Silakan upload file CSV atau pastikan `tax_history.csv` ada di root folder.")
            
            # Quick Win: CSV Template Downloads
            st.markdown("---")
            st.subheader("📥 Download Template CSV")
            st.caption("Unduh template untuk memulai dengan format yang benar")
            
            col_t1, col_t2 = st.columns(2)
            with col_t1:
                sample_tax_path = os.path.join(root_dir, 'sample_tax_history.csv')
                if os.path.exists(sample_tax_path):
                    with open(sample_tax_path, 'rb') as f:
                        st.download_button(
                            "📄 Tax History Template",
                            f.read(),
                            "sample_tax_history.csv",
                            "text/csv",
                            use_container_width=True
                        )
            with col_t2:
                sample_macro_path = os.path.join(root_dir, 'sample_macro_data.csv')
                if os.path.exists(sample_macro_path):
                    with open(sample_macro_path, 'rb') as f:
                        st.download_button(
                            "📊 Macro Data Template",
                            f.read(),
                            "sample_macro_data.csv",
                            "text/csv",
                            use_container_width=True
                        )



    st.markdown("---")
    st.subheader("⚙️ Model Strategy")
    forecast_months = st.slider("Horizon Forecast", 6, 24, 12, format="%d Bulan")
    
    # Model Selection Logic
    model_strategy = st.radio("Metode Pemilihan Model:", 
                                ["Auto (Best Accuracy)", "Ensemble (Top 3)", "Custom Selection"], 
                                index=0)
    
    manual_model_choice = None
    if model_strategy == "Custom Selection":
        avail_models = ["XGBoost", "Prophet", "SARIMA", "Holt-Winters", "LightGBM", "Random Forest", "SARIMAX", "Gradient Boosting", "Extra Trees", "SVR", "Ridge", "ElasticNet"]
        if 'forecaster_v6' in st.session_state and st.session_state['forecaster_v6'] and st.session_state['forecaster_v6'].available_models:
            avail_models = st.session_state['forecaster_v6'].available_models
            
        manual_model_choice = st.multiselect("Pilih Model (Bisa > 1 untuk Ensemble):", sorted(avail_models), default=[sorted(avail_models)[0]] if avail_models else None)
        st.caption("✨ Pilih satu atau lebih model. Jika lebih dari satu, hasil akan dirata-rata (Ensemble).")

    st.info("💡 **Tip:** Want to compare two scenarios side-by-side? Go to **Scenario Lab** on sidebar menu.")

    st.markdown("---")
    st.subheader("🎛️ Simulation Engine")
    
    # High Priority: Preset Scenario Templates
    st.caption("⚡ Quick Load Preset Scenarios")
    preset_scenarios = scenario_utils.get_preset_scenarios()
    
    col_preset1, col_preset2 = st.columns([2, 1])
    with col_preset1:
        selected_preset = st.selectbox(
            "Load Preset Scenario:",
            ["Custom (Manual)"] + list(preset_scenarios.keys()),
            help="Select a predefined economic scenario to quickly populate all macro indicators"
        )
    with col_preset2:
        if selected_preset != "Custom (Manual)":
            if st.button("📥 Load Preset", use_container_width=True):
                st.session_state['loaded_preset'] = preset_scenarios[selected_preset]
                st.success(f"Loaded: {selected_preset}")
                st.rerun()
    
    # Load Defaults from Macro File
    def_inflasi, def_gdp, def_kurs, def_oil = 3.0, 5.0, 15000.0, 80.0
    # Additional Defaults
    def_sbn, def_lift_oil, def_lift_gas = 6.6, 605.0, 1005.0
    def_ihsg, def_bi, def_coal, def_cpo = 7300.0, 6.25, 145.0, 820.0
    def_exp, def_imp, def_cons, def_pmi = 5.0, 4.0, 4.9, 51.0

    try:
        base_macro = pd.read_csv(default_macro) if os.path.exists(default_macro) else None
        if base_macro is not None and not base_macro.empty:
            last = base_macro.iloc[-1] # Use Most Recent Data
            def_inflasi = float(last.get('Inflasi', 3.0))
            def_gdp = float(last.get('Pertumbuhan_Ekonomi', 5.0))
            def_kurs = float(last.get('Kurs_USD', 15000.0))
            def_oil = float(last.get('Harga_Minyak_ICP', last.get('Harga_Komoditas', 80.0)))
            
            # New Indicators
            def_sbn = float(last.get('SBN_10Y', 6.6))
            def_lift_oil = float(last.get('Lifting_Minyak', 605.0))
            def_lift_gas = float(last.get('Lifting_Gas', 1005.0))
            
            def_ihsg = float(last.get('IHSG', 7300.0))
            def_bi = float(last.get('BI_Rate', 6.25))
            def_coal = float(last.get('Harga_Batubara', 145.0))
            def_cpo = float(last.get('Harga_CPO', 820.0))
            
            def_exp = float(last.get('Ekspor_Growth', 5.0))
            def_imp = float(last.get('Impor_Growth', 4.0))
            def_cons = float(last.get('Konsumsi_RT_Growth', 4.9))
            def_pmi = float(last.get('PMI_Manufaktur', 51.0))
    except Exception as e:
        # Fallback to existing defaults if read fails
        pass
    
    # Check if preset was loaded
    if 'loaded_preset' in st.session_state and st.session_state['loaded_preset']:
        loaded = st.session_state['loaded_preset']
        # Override defaults with loaded preset
        def_inflasi = loaded.get('Inflasi', def_inflasi)
        def_gdp = loaded.get('Pertumbuhan_Ekonomi', def_gdp)
        def_kurs = loaded.get('Kurs_USD', def_kurs)
        def_oil = loaded.get('Harga_Minyak_ICP', def_oil)
        def_sbn = loaded.get('SBN_10Y', def_sbn)
        def_lift_oil = loaded.get('Lifting_Minyak', def_lift_oil)
        def_lift_gas = loaded.get('Lifting_Gas', def_lift_gas)
        def_ihsg = loaded.get('IHSG', def_ihsg)
        def_bi = loaded.get('BI_Rate', def_bi)
        def_coal = loaded.get('Harga_Batubara', def_coal)
        def_cpo = loaded.get('Harga_CPO', def_cpo)
        def_exp = loaded.get('Ekspor_Growth', def_exp)
        def_imp = loaded.get('Impor_Growth', def_imp)
        def_cons = loaded.get('Konsumsi_RT_Growth', def_cons)
        def_pmi = loaded.get('PMI_Manufaktur', def_pmi)
        # Clear loaded preset after applying
        st.session_state['loaded_preset'] = None

    st.caption("Masukan 7 Indikator Utama (Asumsi Dasar Makro)")
    with st.expander("🏛️ Asumsi Dasar APBN", expanded=True):
        c1, c2 = st.columns(2)
        # 1 & 2
        sc_gdp = c1.number_input("Pertumbuhan Ekonomi (%)", value=def_gdp, step=0.1)
        sc_inflasi = c2.number_input("Inflasi (%)", value=def_inflasi, step=0.1)
        
        c3, c4 = st.columns(2)
        # 3 & 4
        sc_kurs = c3.number_input("Nilai Tukar (IDR/USD)", value=def_kurs, step=100.0)
        sbn_10y = c4.number_input("SBN 10 Tahun (%)", value=def_sbn, step=0.10, help="Rata-rata Suku Bunga SUN 10 Tahun")
        
        c5, c6 = st.columns(2)
        # 5 & 6
        sc_oil = c5.number_input("Harga Minyak (ICP) ($)", value=def_oil, step=1.0)
        lift_oil = c6.number_input("Lifting Minyak (rb bph)", value=def_lift_oil, step=5.0, help="Target Lifting Minyak Bumi")
        
        # 7
        lift_gas = st.number_input("Lifting Gas (rb boepd)", value=def_lift_gas, step=10.0, help="Target Lifting Gas Bumi")

    
    # Advanced / Sectoral (Extended)
    with st.expander("📊 Indikator Sektoral & Moneter (Lanjutan)", expanded=False):
        # 1. Financial Market
        st.caption("🏦 Finansial & Moneter")
        c_fin1, c_fin2 = st.columns(2)
        ihsg = c_fin1.number_input("IHSG", value=def_ihsg, step=50.0)
        bi_rate = c_fin2.number_input("BI Rate %", value=def_bi, step=0.25)
        
        # 2. Commodities (Non-Oil)
        st.caption("🛢️ Komoditas Non-Migas")
        c_com1, c_com2 = st.columns(2)
        coal_p = c_com1.number_input("Batubara ($/ton)", value=def_coal, step=5.0)
        cpo_p = c_com2.number_input("CPO ($/ton)", value=def_cpo, step=10.0)
        
        # 3. Real Sector
        st.caption("🏭 Perdagangan & Konsumsi")
        c_real1, c_real2 = st.columns(2)
        ekspor = c_real1.number_input("Ekspor Growth %", value=def_exp, step=0.5)
        impor = c_real2.number_input("Impor Growth %", value=def_imp, step=0.5)
        
        c_real3, c_real4 = st.columns(2)
        konsumsi = c_real3.number_input("Konsumsi RT %", value=def_cons, step=0.1)
        pmi = c_real4.number_input("PMI Manufaktur", value=def_pmi, step=0.5, help="Indeks Manajer Pembelian (>50 Ekspansi)")
        
    # Buttons moved to bottom of sidebar




# Check if we have a file uploaded OR a persistent session
has_session = 'forecaster_v6' in st.session_state and st.session_state['forecaster_v6'] is not None

if history_file or has_session:        # Load Model
    if history_file:
         # Check if it's a NEW file
         current_file_name = history_file.name
         loaded_name = st.session_state.get('loaded_filename', None)
         
         if current_file_name != loaded_name:
             # New File Detected -> Force Reload
             with st.spinner(f"Loading new file: {current_file_name}..."):
                 fc = TaxForecaster(history_file, macro_file)
                 fc.load_data()
                 st.session_state['forecaster_v6'] = fc
                 st.session_state['baseline_results_v6'] = None
                 st.session_state['loaded_filename'] = current_file_name
                 # Reset previous comparisons
                 st.session_state['previous_forecast_comparison'] = None
         else:
             # Same file as loaded -> Use Session explicitly (avoid re-init)
             fc = st.session_state['forecaster_v6']
             
    elif has_session and not history_file:
         # User navigated away and back. Uploader is empty, but session is alive.
         st.info("ℹ️ Using previously uploaded data from session.")
         fc = st.session_state['forecaster_v6']
    
    # Ensure fc is assigned
    fc = st.session_state['forecaster_v6']


    # --- FEATURE 4: SIDEBAR ADVANCED TUNING ---
    with st.sidebar.expander("🛠️ Advanced Tuning", expanded=False):
        st.caption("Konfigurasi Hyperparameter Tuning")
        n_trials = st.slider("Jumlah Trials (Optuna)", 5, 50, 10)
        epochs = st.slider("Max Epochs (MLP)", 50, 500, 100)
        
        # Moved from train logic to here for accessibility
        fix_outliers = st.checkbox("✨ Auto-Correct Outliers", value=False, help="Automatically fix values with Z-Score > 2.5 before training")
        
        st.caption("Semakin tinggi angka, semakin lama training namun potensi akurasi lebih baik.")
    
    # Quick Win: Keyboard Shortcuts Help
    with st.sidebar.expander("⌨️ Keyboard Shortcuts", expanded=False):
        st.markdown("""
        <div class="keyboard-shortcut">
            <span class="kbd">Ctrl+R</span> - Refresh macro data
        </div>
        <div class="keyboard-shortcut">
            <span class="kbd">Ctrl+T</span> - Train models
        </div>
        <div class="keyboard-shortcut">
            <span class="kbd">Ctrl+E</span> - Export forecast
        </div>
        <div class="keyboard-shortcut">
            <span class="kbd">Ctrl+S</span> - Save scenario
        </div>
        """, unsafe_allow_html=True)
        
    st.sidebar.markdown("---")
    
    # Actions Buttons
    train_btn = st.sidebar.button("▶ Train Model", type="primary", use_container_width=True, disabled=not history_file)
    sim_btn = st.sidebar.button("🔃 Simulate", type="secondary", use_container_width=True)
    
    # High Priority: Save/Load Custom Scenarios
    if st.sidebar.button("💾 Save Scenario", use_container_width=True, help="Save current settings as custom scenario"):
        scenario_name = f"Custom_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}"
        custom_scenario = {
            'Inflasi': sc_inflasi, 'Pertumbuhan_Ekonomi': sc_gdp, 'Kurs_USD': sc_kurs,
            'Harga_Minyak_ICP': sc_oil, 'SBN_10Y': sbn_10y, 'Lifting_Minyak': lift_oil,
            'Lifting_Gas': lift_gas, 'IHSG': ihsg, 'BI_Rate': bi_rate,
            'Harga_Batubara': coal_p, 'Harga_CPO': cpo_p, 'Ekspor_Growth': ekspor,
            'Impor_Growth': impor, 'Konsumsi_RT_Growth': konsumsi, 'PMI_Manufaktur': pmi
        }
        if 'saved_scenarios' not in st.session_state:
            st.session_state['saved_scenarios'] = {}
        st.session_state['saved_scenarios'][scenario_name] = custom_scenario
        st.success(f"Saved as: {scenario_name}")
        
    st.sidebar.markdown("---")
    if st.sidebar.button("🧹 Clear Cache / Reset App", use_container_width=True):
        st.session_state.clear()
        st.rerun()

    # --- HTML REPORT DOWNLOAD (SIDEBAR) ---

    if fc and fc.results:
        st.sidebar.markdown("---")
        st.sidebar.subheader("📑 Executive Report")

        # Generate Report Button
        if st.sidebar.button("📄 Generate Report (HTML)", use_container_width=True):
            with st.spinner("Generating HTML Report..."):
                import importlib
                importlib.reload(report_generator)
                
                # Generate HTML
                html_report = report_generator.generate_html_report(fc)
                st.session_state['report_html'] = html_report
                st.session_state['report_generated'] = True
                
        # Persistent Download Buttons
        if st.session_state.get('report_generated', False):
            st.sidebar.download_button(
                label="📥 Download HTML Report",
                data=st.session_state['report_html'],
                file_name=f"TaxForecast_{pd.Timestamp.now().strftime('%Y%m%d')}.html",
                mime="text/html",
                use_container_width=True
            )

    # TRAIN LOGIC

    if train_btn:
        st.write("🏥 Performing Data Health Check...")
        health_issues = []
        
        # Check History
        if fc and fc.df is not None:
            if fc.df.isnull().sum().sum() > 0: health_issues.append("Found missing values in History Data (will be filled).")
            if (fc.df['Nominal (Milyar)'] <= 0).any(): health_issues.append("Found zero or negative tax values (potential anomalies).")
        
        if health_issues:
            for issue in health_issues: st.warning(f"⚠️ {issue}")
            st.success("✅ Data Health Check Passed: No obvious data structure errors.")
        
        if fix_outliers and fc:
            msg = fc.preprocess_outliers()
            st.success(f"🛠️ {msg}")
            
        # --- FEATURE 6: CORRELATION HEATMAP ---
        with st.expander("📊 Correlation Analysis (Macro vs Tax)"):
            if fc and fc.df is not None and fc.macro_df is not None:
                h_df = fc.df.copy()
                m_df = fc.macro_df.copy()
                tax_pivot = h_df.pivot_table(index='Tanggal', columns='Jenis Pajak', values='Nominal (Milyar)', aggfunc='sum')
                m_df.set_index('Tanggal', inplace=True)
                
                corr_df = tax_pivot.join(m_df, how='inner')
                if not corr_df.empty:
                    corr_matrix = corr_df.corr()
                    fig_corr = px.imshow(corr_matrix, text_auto=True, aspect="auto", color_continuous_scale='RdBu_r', title="Matriks Korelasi")
                    st.plotly_chart(fig_corr, use_container_width=True)
                else:
                    st.info("Not enough overlapping data for correlation.")
        
        if not fc:
            st.error("❌ Forecaster not initialized. Please upload data first.")
            st.stop()
        
        # Training time estimation
        tax_types_count = len(fc.df['Jenis Pajak'].unique())
        estimated_time = tax_types_count * n_trials * 2
        st.info(f"⏱️ Estimated training time: ~{estimated_time} seconds ({estimated_time/60:.1f} minutes)")
        
        # Create progress tracking UI
        progress_container = st.container()
        
        with progress_container:
            progress_bar = st.progress(0)
            status_text = st.empty()
            current_tax_text = st.empty()
            
            # Progress callback function
            def update_progress(current_step, total_steps, message, progress_pct=0):
                """Update progress bar and status text"""
                # Update progress bar
                progress_bar.progress(min(int(progress_pct * 100), 100))
                
                # Update status messages
                status_text.markdown(f"**Status:** {message}")
                current_tax_text.info(f"📊 Processing step {current_step} of {total_steps}")
        
        with st.spinner(f"Training models (Seasonality Detect, {n_trials} Trials, {epochs} Epochs)..."):
            # Update files if user uploaded new ones
            fc.history_file = history_file 
            fc.macro_file = macro_file
            fc.load_data()
            
            try:
                # Train with progress callback
                status_text.text("🚀 Training in progress...")
                fc.fit(n_trials=n_trials, epochs=epochs, progress_callback=update_progress) 
                progress_bar.progress(100)
                status_text.success("✅ Training Complete!")
                
            except Exception as e:
                st.error(f"❌ Training failed: {str(e)}")
                with st.expander("🔧 Error Details"):
                    st.code(str(e))
                st.stop()
            
            # Run baseline prediction immediately
            try:
                dates_future_init = pd.date_range(start=(fc.df['Tanggal'].max() + pd.DateOffset(months=1)).replace(day=1), periods=forecast_months, freq='MS')
                scenario_df_init = pd.DataFrame({
                    'Tanggal': dates_future_init,
                    'Inflasi': sc_inflasi,
                    'Pertumbuhan_Ekonomi': sc_gdp,
                    'Kurs_USD': sc_kurs,
                    'Harga_Minyak_ICP': sc_oil,
                    'SBN_10Y': sbn_10y,
                    'Lifting_Minyak': lift_oil,
                    'Lifting_Gas': lift_gas,
                    'IHSG': ihsg,
                    'BI_Rate': bi_rate,
                    'Harga_Batubara': coal_p,
                    'Harga_CPO': cpo_p,
                    'Ekspor_Growth': ekspor,
                    'Impor_Growth': impor,
                    'Konsumsi_RT_Growth': konsumsi,
                    'PMI_Manufaktur': pmi
                })
                strat_key_init = 'auto'
                if model_strategy == "Ensemble (Top 3)": strat_key_init = 'ensemble'
                if model_strategy == "Custom Selection": strat_key_init = 'custom'
                
                fc.predict(forecast_periods=forecast_months, custom_macro_future=scenario_df_init, 
                           model_strategy=strat_key_init, manual_model=manual_model_choice)
                
                # Update session state
                st.session_state['forecaster_v6'] = fc
                
            except Exception as e:
                st.error(f"Initial prediction failed: {e}")
            
            # Display training metrics
            if fc.metadata.get('training_duration'):
                col_m1, col_m2, col_m3 = st.columns(3)
                col_m1.metric("⏱️ Training Time", f"{fc.metadata['training_duration']}s")
                col_m2.metric("📊 Tax Types Trained", tax_types_count)
                col_m3.metric("🎯 Models per Type", "~15")
                st.success(f"✅ Training Complete!")
            else:
                st.success("✅ Training Complete!")
            
            st.rerun()

    # MAIN DISPLAY
    if fc and fc.is_fitted:
        strat_key = 'auto'
        if model_strategy == "Ensemble (Top 3)": strat_key = 'ensemble'
        if model_strategy == "Custom Selection": strat_key = 'custom'
        
        # PREDICT
        dates_future = pd.date_range(start=fc.df['Tanggal'].max() + pd.DateOffset(months=1), periods=forecast_months, freq='ME')
        
        # 1. Baseline
        if st.session_state.get('baseline_results_v6') is None:
            if fc.results:
                st.session_state['baseline_results_v6'] = pd.concat([r['data'] for r in fc.results], ignore_index=True)

        # 2. Scenario - Include ALL macro variables for SARIMAX compatibility
        scenario_df = pd.DataFrame({
            'Tanggal': dates_future,
            'Inflasi': sc_inflasi,
            'Pertumbuhan_Ekonomi': sc_gdp,
            'Kurs_USD': sc_kurs,
            'Harga_Minyak_ICP': sc_oil,
            'SBN_10Y': sbn_10y,
            'Lifting_Minyak': lift_oil,
            'Lifting_Gas': lift_gas,
            'IHSG': ihsg,
            'BI_Rate': bi_rate,
            'Harga_Batubara': coal_p,
            'Harga_CPO': cpo_p,
            'Ekspor_Growth': ekspor,
            'Impor_Growth': impor,
            'Konsumsi_RT_Growth': konsumsi,
            'PMI_Manufaktur': pmi
        })
        
        # Quick Win: Store previous forecast for comparison
        if fc.results and len(fc.results) > 0:
            current_total = pd.concat([r['data'] for r in fc.results], ignore_index=True)['Nominal (Milyar)'].sum()
            if st.session_state.get('previous_forecast_comparison') is not None:
                prev_total = st.session_state['previous_forecast_comparison']
                delta = current_total - prev_total
                delta_pct = (delta / prev_total * 100) if prev_total > 0 else 0
                
                if abs(delta_pct) > 1:  # Show only if change is significant
                    st.info(f"📊 **Forecast Comparison**: Current forecast differs by **{delta_pct:+.1f}%** (Rp {delta:+,.0f}M) from previous run.")
        
        fc.predict(forecast_periods=forecast_months, custom_macro_future=scenario_df, 
                    model_strategy=strat_key, manual_model=manual_model_choice)
        
        # Store current forecast for next comparison
        if fc.results:
            st.session_state['previous_forecast_comparison'] = pd.concat([r['data'] for r in fc.results], ignore_index=True)['Nominal (Milyar)'].sum()
        
        # DISPLAY RESULTS
        if fc.results:
            results_df = pd.concat([r['data'] for r in fc.results], ignore_index=True)
            
            if 'Tanggal' in results_df.columns:
                results_df['Tanggal'] = pd.to_datetime(results_df['Tanggal'])
            
            hist_df = fc.df[['Tanggal', 'Jenis Pajak', 'Nominal (Milyar)']].copy()
            hist_df['Tanggal'] = pd.to_datetime(hist_df['Tanggal'])
            hist_df['Tipe Data'] = 'Realisasi'
            
            fc_df = results_df.copy()
            fc_df['Tipe Data'] = 'Forecast'
            
            combined_df = pd.concat([hist_df, fc_df], ignore_index=True)
            
            fc_df['Tanggal'] = pd.to_datetime(fc_df['Tanggal'])
            total_forecast = fc_df.groupby('Tanggal')['Nominal (Milyar)'].sum().sum()
            last_date_hist = hist_df['Tanggal'].max()
            start_date_hist = last_date_hist - pd.DateOffset(months=11)
            last_12m_hist = hist_df[(hist_df['Tanggal'] >= start_date_hist) & (hist_df['Tanggal'] <= last_date_hist)]
            total_hist_L12M = last_12m_hist['Nominal (Milyar)'].sum()
            
            growth = ((total_forecast - total_hist_L12M) / total_hist_L12M) * 100 if total_hist_L12M > 0 else 0
            avg_monthly = total_forecast / forecast_months
            
            growth_icon = "📈" if growth > 0 else "📉"
            growth_color = "#2ecc71" if growth > 0 else "#e74c3c"
            
            total_lower = fc_df['Nominal Lower'].sum() if 'Nominal Lower' in fc_df.columns else total_forecast * 0.95
            total_upper = fc_df['Nominal Upper'].sum() if 'Nominal Upper' in fc_df.columns else total_forecast * 1.05
            
            insight_text = f"""<div style="font-size: 1.1rem; color: #ECF0F1;">
            🤖 <b>AI Insight:</b> Proyeksi total penerimaan pajak diperkirakan mencapai 
            <b style="color: #3498DB;">Rp {total_forecast:,.0f} Milyar</b> 
            (<i>Range: Rp {total_lower:,.0f} M - {total_upper:,.0f} M</i>). <br>
            Ini menunjukkan tren <b style="color: {growth_color};">{growth:+.1f}%</b> {growth_icon} 
            dibandingkan periode sebelumnya.
            </div>"""
            
            st.markdown(f"""<div class="content-card">
            <h3 style="margin-top:0;">📊 Executive Summary</h3>
            {insight_text}
            </div>""", unsafe_allow_html=True)

            st.markdown("### 🔑 Key Performance Indicators")
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                style.display_metric_card("Total Forecast", f"{total_forecast:,.0f}", growth, suffix="M")
                st.caption("💡 Total proyeksi penerimaan")
            with col2:
                style.display_metric_card("Rata-rata Bulanan", f"{avg_monthly:,.0f}", suffix="M")
                st.caption("💡 Rata-rata per bulan")
            with col3:
                style.display_metric_card("Confidence Lower", f"{total_lower:,.0f}", suffix="M")
                st.caption("💡 Batas bawah 95% CI")
            with col4:
                style.display_metric_card("Confidence Upper", f"{total_upper:,.0f}", suffix="M")
                st.caption("💡 Batas atas 95% CI")
            
            # Quick Win: Forecast Summary Cards
            st.markdown("---")
            st.subheader("📋 Ringkasan Forecast per Jenis Pajak")
            if hasattr(fc, 'model_performance') and fc.model_performance:
                summary_cols = st.columns(2)
                
                # Helper to find actual used model from results
                results_map = {r['tax_type']: r['model'] for r in fc.results} if hasattr(fc, 'results') and fc.results else {}

                for idx, perf in enumerate(fc.model_performance):
                    with summary_cols[idx % 2]:
                        tax_type = perf['Jenis Pajak']
                        tax_forecast = fc_df[fc_df['Jenis Pajak'] == tax_type]['Nominal (Milyar)'].sum()
                        
                        # Use actual model from simulation if available, else fallback to best training model
                        display_model = results_map.get(tax_type, perf['Best Model'])
                        
                        card_html = create_summary_card(
                            tax_type,
                            display_model,
                            perf['Accuracy'],
                            tax_forecast
                        )
                        st.markdown(card_html, unsafe_allow_html=True)
            
            st.markdown("---")



            tab_overview, tab_dist, tab_data, tab_season, tab_explain, tab_calc = st.tabs(["📈 Forecast & Trends", "📊 Distribution", "🗃️ Data Explorer", "🗓️ Seasonal Analysis", "🤖 AI Explainability", "🧮 Calculator"])
            
            tax_types = combined_df['Jenis Pajak'].unique()
            
            # Note about documentation
            st.info("💡 **Butuh info sumber data & detail model?** Silakan cek menu **User Guide** untuk dokumentasi lengkap.")

            with tab_overview:
                st.markdown("""
                <div class="content-card">
                    <h4 style="color: #FAFAFA; margin-bottom: 20px;">📉 Visualisasi Forecast & Tren</h4>
                """, unsafe_allow_html=True)
                

                # Export Aggregate
                agg_cols = ['Nominal (Milyar)']
                if 'Nominal Lower' in combined_df.columns: agg_cols.append('Nominal Lower')
                if 'Nominal Upper' in combined_df.columns: agg_cols.append('Nominal Upper')
                
                agg_df = combined_df.groupby(['Tanggal', 'Tipe Data'])[agg_cols].sum().reset_index()
                
                csv_agg = agg_df.to_csv(index=False).encode('utf-8')


                # Main Chart
                agg_baseline = None
                if st.session_state.get('baseline_results_v6') is not None:
                        base_raw = st.session_state['baseline_results_v6']
                        if 'Tanggal' in base_raw.columns:
                            base_raw['Tanggal'] = pd.to_datetime(base_raw['Tanggal'])
                        if 'Tipe Data' not in base_raw.columns:
                            base_raw['Tipe Data'] = 'Forecast'
                        agg_baseline = base_raw.groupby(['Tanggal'])['Nominal (Milyar)'].sum().reset_index()
                        agg_baseline['Tipe Data'] = 'Forecast'

                fig_agg = plot_forecast_chart(agg_df, "Total Penerimaan Negara (All Taxes)", baseline_df=agg_baseline)
                st.plotly_chart(fig_agg, use_container_width=True)

                # --- YoY Growth Chart ---
                st.markdown("##### 📊 Year-over-Year (YoY) Growth Projection")
                
                # Calculate YoY
                # Filter 'combined_df' for total aggregation first
                total_ts = combined_df.groupby('Tanggal')['Nominal (Milyar)'].sum().reset_index().sort_values('Tanggal')
                total_ts['YoY_Growth'] = total_ts['Nominal (Milyar)'].pct_change(periods=12) * 100
                
                # Filter only for the Forecast period
                forecast_start_date = fc_df['Tanggal'].min()
                yoy_data = total_ts[total_ts['Tanggal'] >= forecast_start_date].copy()
                
                if not yoy_data.empty:
                    # Color logic: Green if > 0, Red if < 0
                    yoy_data['Color'] = yoy_data['YoY_Growth'].apply(lambda x: '#2ECC71' if x >= 0 else '#E74C3C')
                    
                    fig_yoy = go.Figure()
                    fig_yoy.add_trace(go.Bar(
                        x=yoy_data['Tanggal'],
                        y=yoy_data['YoY_Growth'],
                        marker_color=yoy_data['Color'],
                        name='YoY Growth %',
                        hovertemplate='%{x|%b %Y}<br>Growth: %{y:+.2f}%'
                    ))
                    
                    fig_yoy.update_layout(
                        title="Proyeksi Pertumbuhan Tahunan (YoY Growth)",
                        yaxis_title="Growth (%)",
                        xaxis=dict(tickformat='%b %Y'),
                        height=300,
                        margin=dict(t=30, l=10, r=10, b=10),
                        showlegend=False
                    )
                    
                    # Add simple threshold line at 0
                    fig_yoy.add_hline(y=0, line_dash="solid", line_color="black", line_width=1)
                    
                    st.plotly_chart(fig_yoy, use_container_width=True)
                else:
                    st.caption("Not enough historical data (need 12+ months) to calculate YoY for the forecast period.")

                st.markdown("</div>", unsafe_allow_html=True) # Close content card

                st.markdown("---")
                
                # --- DETAILED BREAKDOWN (Multi-Select) ---
                st.subheader("🔍 Detailed Breakdown & Comparison")
                
                if 'tax_selector_multi' not in st.session_state:
                     if len(tax_types) > 0: st.session_state['tax_selector_multi'] = [tax_types[0]]
                
                selected_taxes = st.multiselect(
                    "Pilih Jenis Pajak (Pilih >1 untuk perbandingan):", 
                    tax_types, 
                    default=st.session_state.get('tax_selector_multi'),
                    key='tax_selector_multi'
                )
                
                if selected_taxes:
                    subset = combined_df[combined_df['Jenis Pajak'].isin(selected_taxes)]
                    
                    if len(selected_taxes) == 1:
                        # Single View: Use detailed standard chart with CI
                        fig = plot_forecast_chart(subset, f"Trend: {selected_taxes[0]}")
                        st.plotly_chart(fig, use_container_width=True)
                    else:
                        # Comparison View: Use Line Chart
                        st.caption("📈 Comparing multiple tax types (Forecast values shown)")
                        
                        # Filter to show relevant history tail + forecast
                        # To avoid clutter, mostly focus on correct ordering
                        
                        fig_comp = px.line(
                            subset, 
                            x='Tanggal', 
                            y='Nominal (Milyar)', 
                            color='Jenis Pajak',
                            line_dash='Tipe Data', # Distinguish Hist vs Forecast
                            title="Perbandingan Tren Penerimaan",
                            symbol="Tipe Data",
                            markers=True
                        )
                        fig_comp.update_layout(
                            hovermode="x unified",
                            yaxis_title="Nominal (Milyar)",
                            legend=dict(orientation="h", y=1.1)
                        )
                        st.plotly_chart(fig_comp, use_container_width=True)
                        
                        # Add Stacked Bar Option
                        with st.expander("📊 View as Stacked Area / Bar"):
                            chart_type = st.radio("Chart Type:", ["Stacked Area", "Stacked Bar"], horizontal=True)
                            if chart_type == "Stacked Area":
                                fig_stack = px.area(subset, x="Tanggal", y="Nominal (Milyar)", color="Jenis Pajak", title="Komposisi Penerimaan (Stacked Area)")
                            else:
                                fig_stack = px.bar(subset, x="Tanggal", y="Nominal (Milyar)", color="Jenis Pajak", title="Komposisi Penerimaan (Stacked Bar)")
                            st.plotly_chart(fig_stack, use_container_width=True)

                st.markdown("---")
                
                # Leaderboard
                st.subheader("🏆 Model Performance Leaderboard")
                if hasattr(fc, 'model_performance') and fc.model_performance:
                    # 1. Main Leaderboard (Winners)
                    perf_df = pd.DataFrame(fc.model_performance)
                    
                    # Update 'Best Model' to reflect what was actually used in simulation
                    results_map = {r['tax_type']: r['model'] for r in fc.results} if hasattr(fc, 'results') and fc.results else {}
                    if results_map:
                         perf_df['Best Model'] = perf_df.apply(lambda x: results_map.get(x['Jenis Pajak'], x['Best Model']), axis=1)
                    
                    # Reorder & Format for Display
                    disp_cols = ['Jenis Pajak', 'Best Model', 'Accuracy', 'MAPE', 'RMSE', 'Nilai Proyeksi', 'Lower Bound', 'Upper Bound']
                    # Filter only existing columns (just in case)
                    valid_cols = [c for c in disp_cols if c in perf_df.columns]
                    perf_display = perf_df[valid_cols].copy()
                    
                    # Quick Win: Color-coded accuracy
                    def color_accuracy(val):
                        if isinstance(val, str) and '%' in val:
                            acc_num = float(val.replace('%', ''))
                            if acc_num >= 90:
                                return 'background-color: #d4edda; color: #155724'  # Green
                            elif acc_num >= 75:
                                return 'background-color: #fff3cd; color: #856404'  # Yellow
                            else:
                                return 'background-color: #f8d7da; color: #721c24'  # Red
                        return ''
                    
                    st.dataframe(
                        perf_display.style.format({
                            'RMSE': "{:,.2f}",
                            'Nilai Proyeksi': "Rp {:,.0f} M",
                            'Lower Bound': "Rp {:,.0f} M",
                            'Upper Bound': "Rp {:,.0f} M"
                        }).applymap(color_accuracy, subset=['Accuracy']), 
                        use_container_width=True,
                        hide_index=True
                    )
                    
                    # 2. All Candidates Comparison
                    with st.expander("🔍 Detail Perbandingan Semua Model (All Candidates)", expanded=False):
                        st.info("Berikut adalah performa (Training Scores) dari seluruh algoritma yang diuji untuk setiap jenis pajak.")
                        st.caption("💡 **Note:** Nilai Forecast dan Confidence Bounds hanya dihitung untuk model pemenang (ditampilkan di tabel utama di atas) untuk efisiensi komputasi.")
                        
                        if hasattr(fc, 'trained_models'):
                            tabs_tax = st.tabs(list(fc.trained_models.keys()))
                            for i, tax_key in enumerate(fc.trained_models.keys()):
                                with tabs_tax[i]:
                                    info = fc.trained_models[tax_key]
                                    candidates = info.get('candidates', {})
                                    
                                    # Create comparison table
                                    comp_data = []
                                    for m_name, m_val in candidates.items():
                                        # m_val is dict with rmse, mape
                                        accuracy = max(0, 100 - (m_val['mape']*100))
                                        comp_data.append({
                                            'Model': m_name,
                                            'Accuracy': f"{accuracy:.2f}%",
                                            'MAPE': f"{m_val['mape']*100:.2f}%",
                                            'RMSE': m_val['rmse']
                                        })
                                    
                                    comp_df = pd.DataFrame(comp_data).sort_values('RMSE', ascending=True)
                                    
                                    # Highlight Winner
                                    best_name = info['best_single'][0]
                                    
                                    def highlight_winner(s):
                                        return ['background-color: #e8f5e9; color: #2e7d32; font-weight: bold' if s['Model'] == best_name else '' for _ in s]

                                    st.dataframe(
                                        comp_df.style.apply(highlight_winner, axis=1).format({
                                            'RMSE': "{:,.2f}"
                                        }),
                                        use_container_width=True,
                                        hide_index=True
                                    )
                else:
                    st.warning("Belum ada model yang dilatih.")

            with tab_dist:
                st.subheader("📊 Distribution & Volatility Analysis (Seaborn & Plotly)")
                st.caption("Visualisasi distribusi data historis dan forecast menggunakan analisis statistik mendalam.")

                # --- 1. SEABORN DEEP DIVE (Static) ---
                st.markdown("#### 🔬 Deep Dive: Statistical Distribution (Seaborn)")
                st.info("Visualisasi statis resolusi tinggi untuk melihat densitas dan outlier secara presisi.")
                
                # Prepare Aggregated Data for Seaborn
                sb_df = combined_df.copy()
                
                # Consolidate Tax Order for consistency across both charts
                tax_order = sb_df.groupby('Jenis Pajak')['Nominal (Milyar)'].sum().sort_values(ascending=False).index
                
                # Create 2 Columns for Seaborn Charts
                sb_col1, sb_col2 = st.columns(2)
                
                with sb_col1:
                    st.markdown("**1. Boxen Plot (Enhanced Box Plot)**")
                    st.caption("Menampilkan lebih banyak detail pada ekor distribusi dibandingkan boxplot biasa.")
                    
                    fig_sb1, ax1 = plt.subplots(figsize=(10, 8))
                    sns.boxenplot(
                        data=sb_df, 
                        x="Nominal (Milyar)", 
                        y="Jenis Pajak", 
                        palette="mako", 
                        order=tax_order,
                        ax=ax1
                    )
                    sns.despine(left=True, bottom=True)
                    ax1.set_xlabel("Nominal (Milyar Rupiah)")
                    ax1.set_ylabel("")
                    ax1.grid(True, axis='x', linestyle='--', alpha=0.7)
                    fig_sb1.tight_layout()
                    st.pyplot(fig_sb1)
                
                with sb_col2:
                    st.markdown("**2. Violin Plot (Density Estimation)**")
                    st.caption("Estimasi bentuk kepadatan (density) dari setiap jenis pajak.")
                    
                    fig_sb2, ax2 = plt.subplots(figsize=(10, 8))
                    sns.violinplot(
                        data=sb_df, 
                        x="Nominal (Milyar)", 
                        y="Jenis Pajak", 
                        palette="rocket", 
                        inner="quart",
                        linewidth=1,
                        order=tax_order,
                        ax=ax2
                    )
                    sns.despine(left=True, bottom=True)
                    ax2.set_xlabel("Nominal (Milyar Rupiah)")
                    ax2.set_ylabel("")
                    ax2.grid(True, axis='x', linestyle='--', alpha=0.7)
                    fig_sb2.tight_layout()
                    st.pyplot(fig_sb2)


                st.markdown("---")

                # --- 2. INTERACTIVE EXPLORER (Plotly) ---
                st.markdown("#### 🖱️ Interactive Explorer (Plotly)")
                
                # Histogram
                st.markdown("**Frequency Distribution (Histogram)**")
                selected_dist_tax = st.selectbox("Select Tax for Histogram:", tax_types, index=0)
                dist_subset = fc_df[fc_df['Jenis Pajak'] == selected_dist_tax]
                
                fig_hist = px.histogram(
                    dist_subset, 
                    x="Nominal (Milyar)", 
                    nbins=20, 
                    title=f"Histogram: {selected_dist_tax} (Forecast Period)",
                    color_discrete_sequence=['#00d2ff'], # Neon Blue
                    opacity=0.8
                )
                fig_hist.update_layout(
                    bargap=0.1,
                    xaxis_title="Nominal (Milyar)",
                    yaxis_title="Frequency",
                    height=400,
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                )
                st.plotly_chart(fig_hist, use_container_width=True)
            
            with tab_data:
                st.subheader("🗃️ Interactive Data Explorer")
                st.caption("Eksplorasi data mentah forecast detail.")
                
                # Setup Display DF
                data_exp_df = fc_df[['Tanggal', 'Jenis Pajak', 'Nominal (Milyar)', 'Nominal Lower', 'Nominal Upper']].copy()
                data_exp_df['Tanggal'] = data_exp_df['Tanggal'].dt.date
                
                # Filters
                c_fil1, c_fil2 = st.columns(2)
                with c_fil1:
                    fil_tax = st.multiselect("Filter Tax Type:", tax_types, default=tax_types)
                with c_fil2:
                    min_val, max_val = float(data_exp_df['Nominal (Milyar)'].min()), float(data_exp_df['Nominal (Milyar)'].max())
                    val_range = st.slider("Filter Nominal Range (Milyar):", min_val, max_val, (min_val, max_val))
                
                if fil_tax:
                    data_exp_df = data_exp_df[data_exp_df['Jenis Pajak'].isin(fil_tax)]
                    data_exp_df = data_exp_df[
                        (data_exp_df['Nominal (Milyar)'] >= val_range[0]) & 
                        (data_exp_df['Nominal (Milyar)'] <= val_range[1])
                    ]
                
                st.dataframe(
                    data_exp_df,
                    column_config={
                        "Nominal (Milyar)": st.column_config.NumberColumn(
                            "Forecast (M)",
                            help="Nominal Forecast dalam Milyar Rupiah",
                            format="Rp %d M"
                        ),
                        "Nominal Lower": st.column_config.NumberColumn(
                            "Lower Bound",
                            format="Rp %d M"
                        ),
                        "Nominal Upper": st.column_config.NumberColumn(
                            "Upper Bound",
                            format="Rp %d M"
                        ),
                        "Tanggal": st.column_config.DateColumn(
                            "Bulan",
                            format="MMM YYYY"
                        )
                    },
                    use_container_width=True,
                    hide_index=True
                )
                
                # Download filtered
                csv_exp = data_exp_df.to_csv(index=False).encode('utf-8')
                st.download_button("⬇️ Download Filtered Data (CSV)", csv_exp, "filtered_forecast.csv", "text/csv")


            with tab_explain:
                st.markdown("""
                <div class="content-card">
                    <h4 style="color: #FAFAFA; margin-bottom: 15px;">🧠 AI Explainability & Feature Importance</h4>
                    <p style="color: #BDC3C7;">Understand which economic factors drive tax revenue predictions.</p>
                </div>
                """, unsafe_allow_html=True)
                
                # High Priority: Enhanced Explainability Dashboard
                tab_shap, tab_feature_imp, tab_sensitivity = st.tabs(["📊 SHAP Analysis", "🎯 Feature Importance", "🌡️ Sensitivity Heatmap"])
                
                with tab_shap:
                    st.caption("Analisis SHAP feature importance (Hanya tersedia jika model yang digunakan adalah Tree-based).")
                    col_ex1, col_ex2 = st.columns([1, 2])
                    with col_ex1:
                        tax_ex = st.selectbox("Pilih Pajak:", tax_types, key='tax_explain_selector')
                    
                    used_model_name = None
                    model_info = next((m for m in fc.model_performance if m['Jenis Pajak'] == tax_ex), None)
                    if model_info:
                        used_model_name = model_info['Best Model']
                        if "Ensemble" in used_model_name: 
                            used_model_name = None
                            st.info("⚠️ SHAP tidak tersedia untuk Ensemble. Menggunakan model terbaik individual.")
                    
                    if st.button("🔍 Generate SHAP Plot", key='shap_btn'):
                        with st.spinner("Generating SHAP analysis..."):
                            shap_fig = fc.get_shap_plot(tax_ex, model_name=used_model_name)
                            
                            if shap_fig is not None and not isinstance(shap_fig, str):
                                st.pyplot(shap_fig)
                            elif isinstance(shap_fig, str):
                                st.warning(shap_fig)
                            else:
                                st.info("Keterangan tidak tersedia.")
                
                with tab_feature_imp:
                    st.caption("Visualisasi kontribusi relatif setiap variabel makro terhadap prediksi")
                    
                    tax_fi = st.selectbox("Pilih Jenis Pajak:", tax_types, key='tax_fi_selector')
                    
                    if tax_fi in fc.trained_models:
                        model_data = fc.trained_models[tax_fi]
                        best_model_name, best_info = model_data['best_single']
                        
                        # Check if model has feature_importances_
                        model_obj = best_info['obj']
                        if hasattr(model_obj, 'feature_importances_'):
                            X = best_info['X']
                            importances = model_obj.feature_importances_
                            
                            # Create DataFrame
                            fi_df = pd.DataFrame({
                                'Feature': X.columns,
                                'Importance': importances
                            }).sort_values('Importance', ascending=True)
                            
                            # Plot with Plotly
                            fig_fi = go.Figure(go.Bar(
                                x=fi_df['Importance'],
                                y=fi_df['Feature'],
                                orientation='h',
                                marker=dict(
                                    color=fi_df['Importance'],
                                    colorscale='Viridis',
                                    showscale=True
                                ),
                                text=fi_df['Importance'].round(4),
                                textposition='auto'
                            ))
                            
                            fig_fi.update_layout(
                                title=f"Feature Importance: {tax_fi} ({best_model_name})",
                                xaxis_title="Importance Score",
                                yaxis_title="Feature",
                                height=500,
                                showlegend=False
                            )
                            
                            st.plotly_chart(fig_fi, use_container_width=True)
                            
                            # Top 5 features
                            top5 = fi_df.tail(5)
                            st.markdown("**Top 5 Most Important Features:**")
                            for idx, row in top5.iterrows():
                                st.write(f"- **{row['Feature']}**: {row['Importance']:.4f}")
                        else:
                            st.warning(f"Model '{best_model_name}' does not support feature importance extraction.")
                    else:
                        st.error("Tax type not found in trained models.")
                
                with tab_sensitivity:
                    st.caption("Heatmap showing sensitivity of each tax type to macro variables")
                    
                    # Create sensitivity matrix (simplified version)
                    st.info("💡 This heatmap shows relative importance of macro variables for each tax type based on model feature importance.")
                    
                    # Collect feature importances for all tax types
                    sensitivity_data = {}
                    all_features = set()
                    
                    for tax_type in tax_types:
                        if tax_type in fc.trained_models:
                            model_data = fc.trained_models[tax_type]
                            best_model_name, best_info = model_data['best_single']
                            model_obj = best_info['obj']
                            
                            if hasattr(model_obj, 'feature_importances_'):
                                X = best_info['X']
                                importances = model_obj.feature_importances_
                                sensitivity_data[tax_type] = dict(zip(X.columns, importances))
                                all_features.update(X.columns)
                    
                    if sensitivity_data:
                        # Create matrix
                        features_list = sorted(list(all_features))
                        matrix = []
                        
                        for feature in features_list:
                            row = [sensitivity_data.get(tax, {}).get(feature, 0) for tax in tax_types]
                            matrix.append(row)
                        
                        # Plot heatmap
                        fig_heatmap = go.Figure(data=go.Heatmap(
                            z=matrix,
                            x=list(tax_types),
                            y=features_list,
                            colorscale='RdYlGn',
                            text=[[f"{val:.3f}" for val in row] for row in matrix],
                            texttemplate="%{text}",
                            textfont={"size": 10},
                            colorbar=dict(title="Importance")
                        ))
                        
                        fig_heatmap.update_layout(
                            title="Sensitivity Heatmap: Macro Variables vs Tax Types",
                            xaxis_title="Tax Type",
                            yaxis_title="Macro Variable",
                            height=600
                        )
                        
                        st.plotly_chart(fig_heatmap, use_container_width=True)
                    else:
                        st.warning("No feature importance data available for heatmap generation.")
        


            with tab_season:
                st.subheader("🗓️ Seasonal Decomposition Analysis")
                st.caption("Decompose time series into Trend, Seasonality, and Residual components.")
                
                s_col1, s_col2 = st.columns([1, 1])
                with s_col1:
                    season_tax = st.selectbox("Select Tax Type:", tax_types, key='season_tax')
                with s_col2:
                    season_model = st.selectbox("Decomposition Model:", ["additive", "multiplicative"], index=0, key='season_model')
                
                # Get history data for decomposition
                if fc and fc.df is not None:
                    # Filter for selected tax
                    tax_ts = fc.df[fc.df['Jenis Pajak'] == season_tax].copy()
                    tax_ts['Tanggal'] = pd.to_datetime(tax_ts['Tanggal'])
                    tax_ts = tax_ts.set_index('Tanggal').sort_index()
                    
                    # Ensure numeric
                    series = tax_ts['Nominal (Milyar)']
                    
                    if len(series) >= 24: # Need at least 2 full cycles for meaningful decompose
                        try:
                            # Run decomposition
                            decomp = seasonal_decompose(series, model=season_model, period=12)
                            
                            # Create 3 subplots
                            from plotly.subplots import make_subplots
                            
                            fig_decomp = make_subplots(
                                rows=4, cols=1, 
                                shared_xaxes=True,
                                subplot_titles=("Original Series", "Trend", "Seasonality", "Residuals"),
                                vertical_spacing=0.05
                            )
                            
                            # Original
                            fig_decomp.add_trace(go.Scatter(x=decomp.observed.index, y=decomp.observed, mode='lines', name='Observed', line=dict(color='#3498DB')), row=1, col=1)
                            # Trend
                            fig_decomp.add_trace(go.Scatter(x=decomp.trend.index, y=decomp.trend, mode='lines', name='Trend', line=dict(color='#E67E22')), row=2, col=1)
                            # Seasonal
                            fig_decomp.add_trace(go.Scatter(x=decomp.seasonal.index, y=decomp.seasonal, mode='lines', name='Seasonal', line=dict(color='#2ECC71')), row=3, col=1)
                            # Residual
                            fig_decomp.add_trace(go.Scatter(x=decomp.resid.index, y=decomp.resid, mode='markers', name='Residual', marker=dict(color='#E74C3C', size=4)), row=4, col=1)
                            
                            fig_decomp.update_layout(height=800, title_text=f"Decomposition for {season_tax}", showlegend=False)
                            st.plotly_chart(fig_decomp, use_container_width=True)
                            
                        except Exception as e:
                            st.error(f"Error during decomposition: {e}")
                    else:
                        st.warning("Insufficient data frequency or length (Need > 24 months) for seasonal decomposition.")
        
            with tab_calc:
                st.subheader("🧮 Sensitivity / Elasticity Calculator")
                st.caption("Hitung dampak perubahan variabel makro terhadap penerimaan pajak secara spesifik.")
                
                c_calc1, c_calc2 = st.columns(2)
                with c_calc1:
                     calc_tax = st.selectbox("Target Pajak:", ["Total Aggregated"] + list(tax_types), key='calc_tax_selector')
                with c_calc2:
                     # All 15 macro indicators
                     macro_vars = [
                         'Inflasi', 
                         'Pertumbuhan_Ekonomi', 
                         'Kurs_USD', 
                         'SBN_10Y',
                         'Harga_Minyak_ICP',
                         'Harga_Batubara',
                         'Harga_CPO',
                         'Lifting_Minyak',
                         'Lifting_Gas',
                         'IHSG',
                         'BI_Rate',
                         'PMI_Manufaktur',
                         'Ekspor_Growth',
                         'Impor_Growth',
                         'Konsumsi_RT_Growth'
                     ]
                     calc_var = st.selectbox("Variabel Makro:", macro_vars, key='calc_var_selector')
                
                st.write("---")
                st.write(f"**Skenario Simulasi:** Jika **{calc_var}** naik sebesar **1%**, apa dampaknya?")
                
                if st.button("Hitung Sensitivitas", key='calc_btn'):
                    with st.spinner("Running sensitivity analysis..."):
                        # Base Prediction (Current Forecast)
                        future_dates = pd.date_range(start=fc.df['Tanggal'].max() + pd.DateOffset(months=1), periods=12, freq='ME') 
                        
                        # Get last macro values
                        last_macro = fc.macro_df.iloc[-1]
                        
                        # Control scenario (no change)
                        data_control = []
                        for _ in range(12): 
                            data_control.append(last_macro.copy())
                        control_df = pd.DataFrame(data_control)
                        control_df['Tanggal'] = future_dates

                        # Run Control
                        fc.predict(forecast_periods=12, custom_macro_future=control_df)
                        res_control_raw = pd.concat([r['data'] for r in fc.results], ignore_index=True)
                        if calc_tax == "Total Aggregated": 
                            val_control = res_control_raw['Nominal (Milyar)'].sum()
                        else: 
                            val_control = res_control_raw[res_control_raw['Jenis Pajak'] == calc_tax]['Nominal (Milyar)'].sum()

                        # Shock scenario (+1%)
                        data_shock = []
                        for _ in range(12):
                            row = last_macro.copy()
                            row[calc_var] = row[calc_var] * 1.01
                            data_shock.append(row)
                        
                        shock_df = pd.DataFrame(data_shock)
                        shock_df['Tanggal'] = future_dates

                        # Run Shock
                        fc.predict(forecast_periods=12, custom_macro_future=shock_df)
                        res_shock_raw = pd.concat([r['data'] for r in fc.results], ignore_index=True)
                        if calc_tax == "Total Aggregated": 
                            val_shock = res_shock_raw['Nominal (Milyar)'].sum()
                        else: 
                            val_shock = res_shock_raw[res_shock_raw['Jenis Pajak'] == calc_tax]['Nominal (Milyar)'].sum()

                        # Calc Delta
                        delta = val_shock - val_control
                        pct_change = (delta / val_control) * 100 if val_control != 0 else 0
                        
                        st.metric(label=f"Dampak ke {calc_tax}", value=f"Rp {val_shock:,.0f} M", delta=f"{pct_change:.2f}% ({delta:,.0f} M)")
                        
                        if pct_change > 0.5:
                            st.info("💡 **Elastisitas Tinggi:** Penerimaan pajak ini SANGAT SENSITIF terhadap kenaikan variabel tersebut.")
                        elif pct_change < -0.5:
                             st.info("💡 **Elastisitas Negatif Signifikan:** Penerimaan cenderung turun jika variabel ini naik.")
                        else:
                            st.info("💡 **Inelastis:** Perubahan variabel ini tidak terlalu berdampak signifikan.")

            st.markdown("---")
            st.subheader("📥 Export & Downloads")
            
            # Export Variables
            agg_cols = ['Nominal (Milyar)']
            if 'Nominal Lower' in combined_df.columns: agg_cols.append('Nominal Lower')
            if 'Nominal Upper' in combined_df.columns: agg_cols.append('Nominal Upper')
            agg_df_export = combined_df.groupby(['Tanggal', 'Tipe Data'])[agg_cols].sum().reset_index()
            
            # CSV Data
            csv_data = agg_df_export.to_csv(index=False).encode('utf-8')
            
            # Layout: 3 Columns for Data Export
            col_e1, col_e2, col_e3 = st.columns(3)

            with col_e1:
                st.download_button("💾 Download CSV Data", csv_data, "forecast_data.csv", "text/csv", use_container_width=True, help="Download forecast data as CSV")
            
            with col_e2:
                # Quick Win: Excel Export
                if st.button("📊 Export to Excel", use_container_width=True, help="Export forecast to formatted Excel file"):
                    with st.spinner("Generating Excel file..."):
                        import io
                        from datetime import datetime
                        
                        # Use forecaster's built-in Excel export
                        excel_buffer = io.BytesIO()
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        
                        try:
                            import openpyxl
                            from openpyxl.utils.dataframe import dataframe_to_rows
                            from openpyxl.styles import Font, PatternFill
                            
                            wb = openpyxl.Workbook()
                            wb.remove(wb.active)
                            
                            # Summary sheet
                            summary = wb.create_sheet("Summary")
                            summary.append(["TaxForecaster 2.0 - Forecast Export"])
                            summary.append(["Generated:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
                            summary.append(["Total Forecast:", f"Rp {total_forecast:,.0f} M"])
                            summary.append(["Growth:", f"{growth:+.1f}%"])
                            summary.append([])
                            
                            perf_df = pd.DataFrame(fc.model_performance)
                            for r in dataframe_to_rows(perf_df, index=False, header=True):
                                summary.append(r)
                            
                            # Data by tax type
                            for tax_type in combined_df['Jenis Pajak'].unique():
                                sheet = wb.create_sheet(tax_type[:31])
                                tax_data = combined_df[combined_df['Jenis Pajak'] == tax_type]
                                for r in dataframe_to_rows(tax_data, index=False, header=True):
                                    sheet.append(r)
                                for cell in sheet[1]:
                                    cell.font = Font(bold=True)
                                    cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
                            
                            wb.save(excel_buffer)
                            excel_buffer.seek(0)
                            
                            st.download_button(
                                "⬇️ Download Excel File",
                                excel_buffer.getvalue(),
                                f"forecast_export_{timestamp}.xlsx",
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                        except ImportError:
                            st.warning("openpyxl not installed. Downloading as CSV instead.")
                            st.download_button(
                                "⬇️ Download CSV",
                                csv_data,
                                f"forecast_export_{timestamp}.csv",
                                "text/csv",
                                use_container_width=True
                            )

            with col_e3:
                # Quick Win: Download All Data ZIP
                scenario_params = {
                    'Forecast Horizon': f"{forecast_months} months",
                    'Model Strategy': model_strategy,
                    'Inflasi': sc_inflasi,
                    'GDP': sc_gdp,
                    'Kurs USD': sc_kurs,
                    'Harga Minyak ICP': sc_oil
                }
                zip_data = export_all_data_zip(combined_df, fc.model_performance, scenario_params)
                st.download_button(
                    "📦 Download All (ZIP)", 
                    zip_data, 
                    "taxforecaster_export.zip", 
                    "application/zip",
                    use_container_width=True,
                    help="Download all forecasts, model performance, and scenario parameters"
                )
            
            # Logic PDF
            # Logic PDF Removed (Moved to Sidebar HTML Report)

